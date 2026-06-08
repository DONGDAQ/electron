"""叠纸三项目自动报价：X3恋与深空、闪暖、X6ニキ新作
   批次驱动，整页复制在线表重复匹配率数据到模板"""
from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from quote_system.feishu_client import FeishuClient, excel_date_serial_to_date
from quote_system.projects import PROJECTS, resolve_project
from quote_system.save_path_config import get_save_path
from quote_system.paths import get_quote_history_dir

# Default starting batch numbers (used when no history found)
DEFAULT_START_BATCH: dict[str, int] = {
    "liandishenkong": 41,
    "shining_nikki": 43,
    "niki_xinzuo": 105,
}

# Template cell/layout config per project
TEMPLATE_LAYOUT: dict[str, dict] = {
    "liandishenkong": {
        "quote_date_cell": ("本地化报价", "H7"),
        "delivery_date_cell": ("本地化报价", "E39"),
        "price_col": "G",
        "data_row_range": (10, 35),
        "has_tep": True,
        "tep_col": "I",
    },
    "shining_nikki": {
        "quote_date_cell": ("本地化报价", "H7"),
        "delivery_date_cell": ("本地化报价", "E49"),
        "price_col": "G",
        "data_row_range": (10, 45),
        "has_tep": False,
    },
    "niki_xinzuo": {
        "quote_date_cell": ("本地化报价", "I7"),
        "delivery_date_cell": ("本地化报价", "E51"),
        "price_col": "H",
        "data_row_range": (10, 47),
        "has_tep": False,
        "service_col": "F",
    },
}

# Regex patterns to extract batch number from history filenames
BATCH_FILENAME_PATTERNS = [
    re.compile(r'第?(\d+)批[次\-]'),
    re.compile(r'批次(\d+)'),
    re.compile(r'^(\d+)批'),
]


def _extract_batch_from_filename(name: str) -> int | None:
    """从文件名中提取批次号"""
    for pat in BATCH_FILENAME_PATTERNS:
        m = pat.search(name)
        if m:
            return int(m.group(1))
    return None


def _extract_batch_from_sheetname(name: str) -> int | None:
    """从sheet名（如'41批次_重复匹配率'）中提取批次号"""
    m = re.search(r'(\d+)批次_重复匹配率', name)
    return int(m.group(1)) if m else None


def _batch_output_filename(template_name: str, new_batch: int) -> str:
    """根据模板文件名生成输出文件名（替换批次号）"""
    name = template_name.replace("模板.", ".")
    return re.sub(r'\d+', str(new_batch), name, count=1)


def _col_letter_to_idx(col: str) -> int:
    """A->0, B->1, ..."""
    return ord(col.upper()) - 65


def _excel_serial_to_date(serial) -> date | None:
    """Excel日期序列号 → date"""
    try:
        num = float(serial)
    except (TypeError, ValueError):
        return None
    from datetime import timedelta
    base = datetime(1899, 12, 30)
    try:
        return (base + timedelta(days=num)).date()
    except Exception:
        return None


# ── Public API ──────────────────────────────────────────────────────

def detect_batches(project_key: str) -> dict:
    """检测已报价批次和待报价批次

    Returns:
        {"last_batch": int|None, "next_batch": int|None, "available_batches": list[int]}
    """
    project = resolve_project(project_key)
    history_dir = get_quote_history_dir() / project.history_dir

    # 1) 从历史文件夹检测已报价批次（含已结算目录）
    quoted_batches: set[int] = set()
    if history_dir.exists():
        xlsx_paths = list(history_dir.glob("*.xlsx"))
        settled_dir = history_dir / "已结算"
        if settled_dir.exists():
            xlsx_paths += list(settled_dir.rglob("*.xlsx"))
        for f in xlsx_paths:
            b = _extract_batch_from_filename(f.name)
            if b:
                quoted_batches.add(b)

    last_batch = max(quoted_batches) if quoted_batches else None

    # 2) 从在线表检测所有可用的批次
    client = FeishuClient(
        sheet_id=project.sheet_id,
        spreadsheet_token=project.spreadsheet_token,
    )
    available_batches: list[int] = []
    try:
        meta = client._api("GET",
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{project.spreadsheet_token}/metainfo")
        if meta.get("code") == 0:
            for s in meta["data"]["sheets"]:
                title = s.get("title", "")
                b = _extract_batch_from_sheetname(title)
                if b:
                    available_batches.append(b)
    except Exception:
        pass

    available_batches.sort()

    # 3) 确定下一个待报价批次
    next_batch = None
    default_start = DEFAULT_START_BATCH.get(project_key)

    if available_batches:
        if last_batch is not None:
            # 本地有记录：取大于已报价的最大值的最小可用批次
            candidates = [b for b in available_batches if b > last_batch]
            if candidates:
                next_batch = candidates[0]
        elif default_start is not None:
            # 本地无记录：从默认起始批次开始，取大于等于它的最小可用批次
            candidates = [b for b in available_batches if b >= default_start]
            if candidates:
                next_batch = candidates[0]
    else:
        # 在线表无可选批次：如果无历史记录，使用默认起始批次
        if last_batch is None and default_start is not None:
            next_batch = default_start

    return {
        "last_batch": last_batch,
        "next_batch": next_batch,
        "available_batches": available_batches,
    }


def get_batch_sheet_id(project_key: str, batch_no: int) -> str | None:
    """获取在线表中对应批次的重复匹配率 sheetId"""
    project = resolve_project(project_key)
    client = FeishuClient(
        sheet_id=project.sheet_id,
        spreadsheet_token=project.spreadsheet_token,
    )
    meta = client._api("GET",
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{project.spreadsheet_token}/metainfo")
    if meta.get("code") != 0:
        return None
    for s in meta["data"]["sheets"]:
        b = _extract_batch_from_sheetname(s.get("title", ""))
        if b == batch_no:
            return s["sheetId"]
    return None


X3_PRICE_RULES = {
    ("TEP", ""): (0.52, "TEP"),
    ("润色/リライト", "非我方初翻"): (0.4, "润色/非我方初翻"),
    ("润色/リライト", "我方初翻"): (0.26, "润色/我方初翻"),
    ("TEP,润色/リライト", "我方初翻"): (0.78, "TEP润色/我方初翻"),
}


def _get_batch_file_service_info(project_key: str, batch_no: int) -> list[dict]:
    """从排期表读取批次中每个文件的E列(服务类型)和F列(初翻归属)"""
    project = resolve_project(project_key)
    client = FeishuClient(
        sheet_id=project.sheet_id,
        spreadsheet_token=project.spreadsheet_token,
    )
    try:
        rows = client.read_sheet()
    except Exception:
        return []

    bcol = 0  # A列: 批次号
    ccol = 2  # C列: 文件名
    ecol = 4  # E列: 服务类型
    fcol = 5  # F列: 初翻归属

    files = []
    last_batch_val = None
    for i, row in enumerate(rows):
        if i == 0:
            continue
        batch_val = row[bcol] if len(row) > bcol else None
        if batch_val is not None and str(batch_val).strip():
            last_batch_val = str(batch_val).strip()
        try:
            row_batch = int(last_batch_val) if last_batch_val is not None else None
        except (ValueError, TypeError):
            continue
        if row_batch != batch_no:
            continue

        c_val = str(row[ccol] or "").strip() if len(row) > ccol else ""
        e_val = str(row[ecol] or "").strip() if len(row) > ecol else ""
        f_val = str(row[fcol] or "").strip() if len(row) > fcol else ""
        if c_val:
            files.append({"name": c_val.split(".")[0], "e": e_val, "f": f_val})

    return files


def _match_file_service(batch_file_name: str, service_info: list[dict]) -> dict:
    """根据批次数据中的文件名匹配排期表的E/F值"""
    # 清理批次文件名：去掉引号、路径、扩展名
    clean = batch_file_name.replace('"', '').strip()
    base = clean.rsplit(".", 1)[0] if "." in clean else clean
    base = base.rsplit("\\", 1)[-1].rsplit("/", 1)[-1]

    for info in service_info:
        sched_name = info["name"]
        # 双向包含匹配（至少4个字符避免误匹配）
        if (len(sched_name) > 3 and sched_name in base) or \
           (len(base) > 3 and base in sched_name):
            return info
    return {"e": "TEP", "f": ""}


def _x3_price_and_note(e_val: str, f_val: str) -> tuple[float, str]:
    """根据E列F列返回(单价, 备注)"""
    for (e, f), (price, note) in X3_PRICE_RULES.items():
        if e_val == e and f_val == f:
            return price, note
    return 0.52, "TEP"


def get_delivery_date(project_key: str, batch_no: int) -> date | None:
    """从排期表中找到该批次最晚的交付日期"""
    project = resolve_project(project_key)
    client = FeishuClient(
        sheet_id=project.sheet_id,
        spreadsheet_token=project.spreadsheet_token,
    )

    # 读取排期表全部数据
    try:
        rows = client.read_sheet()
    except Exception:
        return None

    # 各项目批次列和交付日期列索引（0-based）
    batch_col_map = {
        "liandishenkong": 0,   # 需求批次
        "shining_nikki": 2,    # 批次
        "niki_xinzuo": 4,      # 批次
    }
    deliv_col_map = {
        "liandishenkong": 7,   # 翻译交付
        "shining_nikki": 7,    # 预估纳品日期
        "niki_xinzuo": 9,      # 预估纳品日期
    }

    bcol = batch_col_map.get(project_key)
    dcol = deliv_col_map.get(project_key)
    if bcol is None or dcol is None:
        return None

    latest: date | None = None
    last_batch_val = None
    for i, row in enumerate(rows):
        if i == 0:
            continue
        batch_val = row[bcol] if len(row) > bcol else None
        # 处理合并单元格：将上一个非空值向前传递
        if batch_val is not None and str(batch_val).strip():
            last_batch_val = str(batch_val).strip()
        try:
            row_batch = int(last_batch_val) if last_batch_val is not None else None
        except (ValueError, TypeError):
            continue
        if row_batch != batch_no:
            continue

        dval = row[dcol] if len(row) > dcol else None
        d = _excel_serial_to_date(dval)
        if d and (latest is None or d > latest):
            latest = d

    return latest


def read_batch_sheet_data(project_key: str, sheet_id: str) -> list[list[Any]]:
    """读取在线表某页的全部数据（整页复制用）"""
    project = resolve_project(project_key)
    client = FeishuClient(
        sheet_id=project.sheet_id,
        spreadsheet_token=project.spreadsheet_token,
    )
    return client.read_sheet(sheet_id=sheet_id)


def generate_batch_quote(
    project_key: str,
    batch_no: int,
    quote_date: date,
    price: float = 0.32,
    service_type: str = "初翻",
) -> Path:
    """生成批次报价单

    Returns:
        输出文件路径
    """
    project = resolve_project(project_key)
    layout = TEMPLATE_LAYOUT.get(project_key)
    if not layout:
        raise ValueError(f"未知项目布局配置: {project_key}")

    history_dir = get_quote_history_dir() / project.history_dir
    history_dir.mkdir(parents=True, exist_ok=True)

    template_path = ROOT / "模板" / "报价单模板" / project.template_file
    if not template_path.exists():
        raise FileNotFoundError(f"模板不存在: {template_path}")

    # 确定输出文件名
    out_name = _batch_output_filename(project.template_file, batch_no)
    out_path = history_dir / out_name
    if out_path.exists():
        # 不覆盖，加后缀
        stem = out_path.stem
        for idx in range(2, 1000):
            candidate = out_path.with_name(f"{stem}_{idx}{out_path.suffix}")
            if not candidate.exists():
                out_path = candidate
                break

    # 1) 获取在线表的批次sheet数据
    batch_sheet_id = get_batch_sheet_id(project_key, batch_no)
    if not batch_sheet_id:
        raise ValueError(f"在线表中未找到批次 {batch_no} 的重复匹配率页")

    online_data = read_batch_sheet_data(project_key, batch_sheet_id)

    # 2) 获取交付日期
    delivery_date = get_delivery_date(project_key, batch_no)

    # 3) 复制模板，处理
    shutil.copy2(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    # 4) 清空重复匹配率sheet并粘贴数据
    rate_ws = wb["重复匹配率"]
    # 清空所有已有内容
    max_row = rate_ws.max_row
    max_col = rate_ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            rate_ws.cell(r, c).value = None

    # 写入在线数据
    for r_idx, row_data in enumerate(online_data, start=1):
        for c_idx, cell_val in enumerate(row_data, start=1):
            rate_ws.cell(r_idx, c_idx).value = cell_val

    # 5) 设置报价日期（格式：5月15日）
    qd_sheet, qd_cell = layout["quote_date_cell"]
    qd_ws = wb[qd_sheet]
    qd_ws[qd_cell] = f"{quote_date.month}月{quote_date.day}日"

    # 更新报价单号（有些模板用公式=DATE，有些硬编码了）
    # 如果有公式则不需要手动更新，但H7/ I7是日期单元格

    # 6) 设置交付日期（如果模板有对应单元格）
    if layout.get("delivery_date_cell") and delivery_date:
        dd_sheet, dd_cell = layout["delivery_date_cell"]
        dd_ws = wb[dd_sheet]
        dd_ws[dd_cell] = datetime.combine(delivery_date, datetime.min.time())

    # 7) 设置单价和备注
    price_col = layout["price_col"]
    price_col_idx = _col_letter_to_idx(price_col) + 1  # 1-based for openpyxl
    start_row, end_row = layout["data_row_range"]
    quote_ws = wb["本地化报价"]

    if project_key == "liandishenkong":
        # X3: 根据排期表E列F列设置每行单价和备注
        service_info = _get_batch_file_service_info(project_key, batch_no)
        note_col_idx = _col_letter_to_idx(layout.get("tep_col", "I")) + 1
        file_count = max(0, -(-(len(online_data) - 20) // 20))  # ceiling division
        for r in range(start_row, end_row + 1):
            file_idx = r - start_row  # 0=文件1, 1=文件2, ...
            if file_idx < file_count:
                # 文件名在每个20行块的第2行：22, 42, 62, ...
                file_name_row = 21 + file_idx * 20
                file_name = str(online_data[file_name_row][1] or "") if len(online_data) > file_name_row else ""
                si = _match_file_service(file_name, service_info)
            else:
                continue
            p, n = _x3_price_and_note(si["e"], si["f"])
            quote_ws.cell(r, price_col_idx).value = p
            quote_ws.cell(r, note_col_idx).value = n
        # 合计行（第36行）清空单价和备注
        total_row = end_row + 1
        quote_ws.cell(total_row, price_col_idx).value = None
        quote_ws.cell(total_row, note_col_idx).value = None
    else:
        for r in range(start_row, end_row + 1):
            quote_ws.cell(r, price_col_idx).value = price

        # 8) X6ニキ新作: 填服务类型（初翻/初翻+审校）
        if layout.get("has_tep") and layout.get("tep_col"):
            tep_col_idx = _col_letter_to_idx(layout["tep_col"]) + 1
            for r in range(start_row, end_row + 1):
                quote_ws.cell(r, tep_col_idx).value = "TEP"

    # 9) X6ニキ新作: 填服务类型（初翻/初翻+审校）
    if layout.get("service_col"):
        svc_col_idx = _col_letter_to_idx(layout["service_col"]) + 1
        for r in range(start_row, end_row + 1):
            quote_ws.cell(r, svc_col_idx).value = service_type

    # 10) 隐藏#VALUE!行（数据块不存在的行 = 字数=0 = #VALUE!）
    data_rows = len(online_data)
    # 前20行是所有文件的汇总，每个文件占20行
    file_data_rows = max(0, data_rows - 20)
    file_count = -(-file_data_rows // 20)  # ceiling division
    for r in range(start_row, end_row + 1):
        file_idx = r - start_row  # 0=文件1, 1=文件2, ...
        if file_idx >= file_count:
            quote_ws.row_dimensions[r].hidden = True

    # 11) 启用自动计算
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    wb.save(out_path)
    wb.close()

    save_path = get_save_path(project_key)
    if save_path:
        dest = Path(save_path) / out_path.name
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(out_path), str(dest))

    return out_path


def run_for_project(project_key: str, price: float | None = None,
                    service_type: str = "初翻") -> dict:
    """一键运行：检测批次 → 生成报价单"""
    info = detect_batches(project_key)
    next_batch = info["next_batch"]
    if next_batch is None:
        return {"status": "error", "message": "没有待报价的批次"}

    project = resolve_project(project_key)

    # 如果没有指定价格，使用项目默认第一个语种的价格
    if price is None:
        price = list(project.prices.values())[0] if project.prices else 0.32

    out_path = generate_batch_quote(
        project_key=project_key,
        batch_no=next_batch,
        quote_date=date.today(),
        price=price,
        service_type=service_type,
    )

    return {
        "status": "success",
        "batch": next_batch,
        "output_path": str(out_path),
        "output_name": out_path.name,
    }


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    for pk in ["liandishenkong", "shining_nikki", "niki_xinzuo"]:
        print(f"\n=== {pk} ===")
        info = detect_batches(pk)
        print(f"  已报价批次: {info['last_batch']}")
        print(f"  待报价批次: {info['next_batch']}")
        print(f"  在线表可用批次: {info['available_batches']}")

        if info["next_batch"]:
            try:
                result = run_for_project(pk)
                print(f"  [OK] 已生成: {result['output_name']}")
            except Exception as e:
                print(f"  [FAIL] 失败: {e}")
        else:
            print(f"  - 跳过（无待报价批次）")
