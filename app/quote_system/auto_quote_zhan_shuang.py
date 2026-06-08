"""战双自动报价：从飞书表格读取邮件来源需求 → 下载HTML → 生成报价单"""
from __future__ import annotations

import sys
import traceback
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from quote_system.feishu_client import FeishuClient, excel_date_serial_to_date
from quote_system.projects import resolve_project
from quote_system.generator import QuoteRequest, generate_quote
from quote_system.memoq_html import parse_memoq_html
from quote_system.save_path_config import get_save_path
from quote_system.paths import get_quote_history_dir

SPREADSHEET_TOKEN = "Wup0wnUPIiiIr2k8T23c4zASnjd"
SHEET_ID = "WmpiFq"

# 语种映射：表格中的语种 → 系统用的语种名
LANG_MAP = {
    "中-韩": "中译韩",
    "中译韩": "中译韩",
    "中-英": "中译英",
    "中译英": "中译英",
    "英-中": "英译中",
    "英译中": "英译中",
}


def run(project_key: str = "zhan_shuang") -> None:
    project = resolve_project(project_key)
    work_dir = get_quote_history_dir() / project.history_dir

    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {project.display_name}报价自动化开始")

    client = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN, sheet_id=SHEET_ID)

    try:
        rows = client.read_sheet()
    except Exception as e:
        print(f"读取飞书表格失败: {e}")
        return

    unquoted = find_mail_rows(rows)
    if not unquoted:
        print("没有找到待报价的邮件需求")
        return

    print(f"找到 {len(unquoted)} 个待报价需求")

    # 按 A列(需求名) 分组，同名行合并为一份报价单（填写多行）
    groups: dict[str, list[dict]] = {}
    for item in unquoted:
        key = item["req_name"]
        groups.setdefault(key, []).append(item)

    for req_name, items_in_group in groups.items():
        print(f"\n组 [{req_name}]: {len(items_in_group)} 行")
        for item in items_in_group:
            print(f"  行{item['row_num']}: 语种={item['lang']} | 日期={item['date_label']}")

    for req_name, items_in_group in groups.items():
        try:
            process_group(client, project, work_dir, items_in_group)
            for item in items_in_group:
                client.write_cell(item["row_num"], "F", "已生成")
            print(f"  组 [{req_name}] 状态已更新为'已生成'")
        except Exception as e:
            print(f"处理组 [{req_name}] 失败: {e}")
            traceback.print_exc()


def find_mail_rows(rows: list[list]) -> list[dict]:
    """找出邮件来源、状态为空、有附件的行。"""
    result = []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        row_num = i + 1

        source = str(row[1]).strip() if len(row) > 1 else ""
        if source != "邮件":
            continue

        status = str(row[5]).strip() if len(row) > 5 else ""
        if status and status != "None":
            continue

        name = str(row[0]).strip() if len(row) > 0 else ""
        lang = str(row[4]).strip() if len(row) > 4 else ""
        g_val = row[6] if len(row) > 6 else None

        if not g_val or not isinstance(g_val, list) or len(g_val) == 0:
            continue

        attach = g_val[0]
        file_token = attach.get("fileToken")
        file_name = attach.get("text", "")

        if not file_token:
            continue

        date_serial = row[2] if len(row) > 2 else None
        date_str = excel_date_serial_to_date(date_serial) if date_serial else ""

        deliv_serial = row[3] if len(row) > 3 else None
        deliv_str = excel_date_serial_to_date(deliv_serial) if deliv_serial else ""

        result.append({
            "row_num": row_num,
            "seq": row_num,
            "req_name": name,
            "lang": lang,
            "file_token": file_token,
            "file_name": file_name,
            "date_str": date_str,
            "deliv_str": deliv_str,
            "date_label": date_str or datetime.now().strftime("%Y%m%d"),
        })

    # 按日期排序
    result.sort(key=lambda x: x["date_str"])
    return result


def calc_billable(stats, lang: str = "") -> float:
    """计算战双报价字数：85%以上匹配不计费，其余全量计费。"""
    zero_rate_types = {"X-translated / double context", "Repetition", "101%",
                       "100%", "95%-99%", "context", "exact"}
    is_en = lang == "英-韩"
    total = 0.0
    for row in stats.rows:
        t = row.type.strip() if row.type else ""
        if t in zero_rate_types or t in ("All",):
            continue
        if is_en:
            total += row.source_non_asian_words or 0
        else:
            total += row.source_asian_characters or 0
    return round(total, 2)


def process_group(client: FeishuClient, project, work_dir: Path, items: list[dict]):
    req_name = items[0]["req_name"]
    print(f"\n--- 处理组 [{req_name}]: {len(items)} 行 ---")

    html_dir = ROOT / "outputs" / "zhan_shuang_html"
    html_dir.mkdir(parents=True, exist_ok=True)

    html_paths = []
    stats_list = []
    billables = []

    for item in items:
        html_path = html_dir / item["file_name"]
        print(f"  下载HTML: {item['file_name']}")
        client.download_attachment(item["file_token"], html_path)
        html_paths.append(html_path)

        stats = parse_memoq_html(html_path)
        stats_list.append(stats)
        all_row = stats.all_row
        print(f"    字数: 亚洲字符={all_row.source_asian_characters}, 总字符={all_row.source_chars}")
        billables.append(calc_billable(stats, item["lang"]))

    try:
        quote_date = date.today()
        if items[0]["date_str"]:
            quote_date = datetime.strptime(items[0]["date_str"], "%Y-%m-%d").date()

        delivery_date = None
        if items[0]["deliv_str"]:
            delivery_date = datetime.strptime(items[0]["deliv_str"], "%Y-%m-%d").date()

        system_lang = LANG_MAP.get(items[0]["lang"], items[0]["lang"])
        print(f"  语种: {items[0]['lang']} → {system_lang}")

        save_path = get_save_path("zhan_shuang")
        output_path = Path(save_path) if save_path else None

        print(f"  生成报价单（{len(items)} 个需求合并）...")
        request = QuoteRequest(
            project=project,
            html_paths=[str(p) for p in html_paths],
            languages=[system_lang],
            quote_date=quote_date,
            delivery_date=delivery_date,
            service_content=None,
            request_name=req_name,
            include_extract=False,
            output_path=output_path,
            task_no=None,
            delivery_time=None,
            date_label=items[0].get("date_label"),
        )
        result = generate_quote(ROOT, request)
        print(f"  报价单已生成: {result.final_path}")

        try:
            from settlement.settlement_tracker import quick_record, add_record
            total_billable = sum(billables)
            price = project.prices.get(system_lang, 0.64)
            add_record(quick_record(
                project_key="zhan_shuang",
                company=project.company or "库洛游戏",
                req_name=req_name,
                word_count=round(total_billable),
                total_price=round(total_billable * price * 1.06, 2),
                quote_file=result.final_path.name,
                language=system_lang,
                delivery_date=delivery_date.strftime("%Y-%m-%d") if delivery_date else None,
                source="auto",
                billable_words=round(total_billable),
                source_chars=sum(s.all_row.source_chars or 0 for s in stats_list),
            ))
        except Exception:
            pass

        # 读取C10作为报价页文件名
        wb = openpyxl.load_workbook(result.final_path, data_only=False)
        ws = wb[wb.sheetnames[0]]
        page_fname = ws["C10"].value or ""
        wb.close()

        # 回写每个行的H列（相同文件名）和I列（各自的报价字数）
        for item, billable in zip(items, billables):
            client.write_cell(item["row_num"], "H", str(page_fname))
            client.write_cell(item["row_num"], "I", billable)
            print(f"  行{item['row_num']} H/I 列已写入: {page_fname} / {billable}")

    finally:
        for html_path in html_paths:
            if html_path.exists():
                html_path.unlink()
        print(f"  已删除 {len(html_paths)} 个临时HTML文件")


if __name__ == "__main__":
    run()
