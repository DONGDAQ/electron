from __future__ import annotations

import io
import os
import shutil
import subprocess
import sys
import threading
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, url_for
from markupsafe import Markup, escape

from .config import get_effective_language_config, save_language_config, reset_to_default
from .generator import QuoteRequest, generate_quote
from .memoq_html import quote_words
from .projects import PROJECTS, resolve_project, get_projects_by_company, save_project_config
from .save_path_config import get_save_path, set_save_path
from .paths import get_quote_history_dir, get_settlement_dir
from settlement.settlement_tracker import record_from_quote, add_record as _add_quote_record
from .auto_quote import run as auto_quote_run
from .auto_quote_zhan_shuang import run as zhan_shuang_auto_quote_run
from .auto_quote_zhan_shuang_feishu import run as zhan_shuang_feishu_run
from .auto_quote_diezhi import detect_batches, generate_batch_quote, run_for_project, get_delivery_date


ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
UPLOAD_ROOT = ROOT / "outputs" / "uploads"

OUTPUTS_DIR = Path(r"D:\baojia\electron\outputs")

app = Flask(__name__)
app.secret_key = "local-quote-system"
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
TK_FILL_PROCESS: subprocess.Popen | None = None
FILL_4399_PROCESS: subprocess.Popen | None = None
LOCAL_ONLY_ENDPOINTS = {"/delete-quote", "/open-file", "/open-folder"}
DELETABLE_SUFFIXES = {".xlsx", ".xlsm", ".xls", ".docx", ".pdf"}


def is_local_request() -> bool:
    return request.remote_addr in {"127.0.0.1", "::1", "localhost"}


@app.before_request
def _protect_local_file_actions():
    if request.path in LOCAL_ONLY_ENDPOINTS and not is_local_request():
        return jsonify({"status": "error", "message": "该操作仅允许在本机执行"}), 403


@app.after_request
def _no_cache(response: Response) -> Response:
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/")
def index() -> str:
    projects_by_company = get_projects_by_company()
    all_projects = []
    for company, projects in projects_by_company.items():
        all_projects.extend(projects)
    
    selected_key = request.args.get("project", all_projects[0].key if all_projects else "")
    selected_project = PROJECTS.get(selected_key, all_projects[0] if all_projects else None)
    
    today = date.today().isoformat()
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    
    languages, default_languages = get_effective_language_config(selected_project) if selected_project else ({}, [])
    
    save_path = get_save_path(selected_key) if selected_key else None

    return render_template(
        "index.html",
        projects_by_company=projects_by_company,
        selected_project=selected_project,
        today=today,
        tomorrow=tomorrow,
        languages=languages,
        default_languages=default_languages,
        save_path=save_path,
    )


@app.post("/generate")
def generate() -> Response:
    try:
        project = resolve_project(request.form["project"])
        html_paths = save_uploads(request.files.getlist("html_files"))
        if not html_paths:
            raise ValueError("请至少上传一个 MemoQ HTML 文件")
        
        per_file_languages = None
        all_languages = set()
        all_prices = {}
        
        if len(html_paths) > 1:
            per_file_languages = []
            all_empty = True
            for file_index in range(len(html_paths)):
                file_lang_list = request.form.getlist(f"file_languages[{file_index}]")
                file_price_list = request.form.getlist(f"file_prices[{file_index}]")
                per_file_languages.append(file_lang_list)
                if file_lang_list:
                    all_empty = False
                for lang, price in zip(file_lang_list, file_price_list):
                    if lang:
                        all_languages.add(lang)
                        try:
                            all_prices[lang] = float(price)
                        except ValueError:
                            pass
            
            if all_empty:
                per_file_languages = None
        
        languages, prices = parse_language_prices(project.key)
        
        if not languages and all_languages:
            languages = list(all_languages)
            prices.update(all_prices)
        elif languages:
            for lang in languages:
                if lang in all_prices:
                    prices[lang] = all_prices[lang]
        
        quote_date = parse_date(request.form.get("quote_date")) or date.today()
        delivery_date = parse_date(request.form.get("delivery_date"))
        request_name = clean_optional(request.form.get("request_name"))
        service_content = clean_optional(request.form.get("service_content"))
        include_extract = "摘字" in languages if languages else False
        
        custom_save_path = clean_optional(request.form.get("save_path"))
        output_path = Path(custom_save_path) if custom_save_path else None

        quote_request = QuoteRequest(
            project=project,
            html_paths=html_paths,
            languages=languages,
            quote_date=quote_date,
            delivery_date=delivery_date,
            service_content=service_content,
            request_name=request_name,
            include_extract=include_extract,
            price_overrides=prices,
            per_file_languages=per_file_languages,
            output_path=output_path,
        )
        result = generate_quote(ROOT, quote_request)

        try:
            rec = record_from_quote(quote_request, result.stats, result.final_path, source="manual")
            _add_quote_record(rec)
        except Exception:
            pass

        summary = summarize_stats(project.key, result)
        display_path = result.final_path
        escaped_path = str(display_path).replace("\\", "\\\\")
        message = Markup("已生成报价单：") + Markup('<span onclick="openFileInExcel(\'{}\')" style="cursor:pointer;text-decoration:underline;color:#10b981;">{}</span>').format(
            escaped_path,
            escape(display_path.name),
        ) + Markup("。") + escape(summary)
        flash(message, "success")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("index", project=request.form.get("project", "")))


@app.get("/mobile")
def mobile() -> str:
    return render_template("mobile.html")


@app.post("/mobile/generate")
def mobile_generate() -> Response:
    try:
        message = request.form.get("message", "").strip()
        files = request.files.getlist("files")
        
        if not files:
            return jsonify({
                "status": "error",
                "message": "请上传 HTML 文件"
            })
        
        project_key = request.form.get("project", "")
        demand_date_str = request.form.get("demand_date", "")
        delivery_date_str = request.form.get("delivery_date", "")
        
        if not project_key:
            project_key = detect_project(message)
        
        if not project_key:
            return jsonify({
                "status": "error",
                "message": "无法识别项目，请告诉我项目名称（如：马娘、战双、Bang2）"
            })
        
        try:
            project = resolve_project(project_key)
        except ValueError:
            return jsonify({
                "status": "error",
                "message": f"未找到项目：{project_key}"
            })
        
        demand_date = parse_date(demand_date_str) if demand_date_str else date.today()
        delivery_date = parse_date(delivery_date_str) if delivery_date_str else None
        
        if not delivery_date:
            dates = extract_dates_from_message(message)
            if dates:
                demand_date = dates[0] if len(dates) >= 1 else demand_date
                delivery_date = dates[1] if len(dates) >= 2 else None
        
        saved_files = save_uploads(files)
        html_paths = saved_files
        
        languages, prices = get_effective_language_config(project.key)
        
        stats = []
        for html_path in html_paths:
            try:
                stat = quote_words(html_path)
                stats.append(stat)
            except Exception as exc:
                print(f"处理文件出错 {html_path}: {exc}")
        
        if not stats:
            return jsonify({
                "status": "error",
                "message": "无法读取 HTML 文件，请确保文件格式正确"
            })
        
        save_path = get_save_path(project.key)
        output_path = Path(save_path) if save_path else None

        quote_request = QuoteRequest(
            project=project,
            html_paths=html_paths,
            languages=languages,
            quote_date=demand_date,
            delivery_date=delivery_date,
            service_content=None,
            request_name=None,
            include_extract="摘字" in languages if languages else False,
            price_overrides=prices,
            per_file_languages=None,
            output_path=output_path,
        )
        
        result = generate_quote(ROOT, quote_request)

        try:
            rec = record_from_quote(quote_request, result.stats, result.final_path, source="mobile")
            _add_quote_record(rec)
        except Exception:
            pass

        download_url = url_for("download", path=result.output_path)
        
        total_chars = sum(stat.all_row.source_chars or 0 for stat in stats)
        total_price = 0
        quote_items = []
        
        if languages:
            for lang in languages:
                price = prices.get(lang, 0)
                if lang == "摘字":
                    words = total_chars
                else:
                    words = sum(stat.all_row.target_chars or 0 for stat in stats)
                subtotal = words * price
                total_price += subtotal
                quote_items.append({
                    "language": lang,
                    "words": words,
                    "price": f"{price:.3f}",
                    "subtotal": f"{subtotal:.2f}"
                })
        
        quote_info = {
            "items": quote_items,
            "total": f"{total_price:.2f}"
        }
        
        return jsonify({
            "status": "success",
            "message": "✅ 报价单已生成！",
            "project": project.key,
            "demand_date": demand_date.isoformat(),
            "delivery_date": delivery_date.isoformat() if delivery_date else None,
            "download_url": download_url,
            "quote_info": quote_info
        })
        
    except Exception as exc:
        print(f"移动端生成报价单错误: {exc}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": f"生成失败：{str(exc)}"
        })


def detect_project(message: str) -> str | None:
    message_lower = message.lower()
    
    for proj in PROJECTS:
        for alias in proj.aliases:
            if alias.lower() in message_lower:
                return proj.key
    
    if "马娘" in message or "优俊" in message:
        return "umamusume"
    if "战双" in message:
        return "zhan_shuang"
    if "bang" in message_lower or "bang2" in message_lower:
        return "bang2"
    
    return None


def extract_dates_from_message(message: str) -> list[date]:
    import re
    dates = []
    
    date_patterns = [
        r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})",
        r"(\d{1,2})[月/-](\d{1,2})[日]?",
    ]
    
    for pattern in date_patterns:
        matches = re.findall(pattern, message)
        for match in matches:
            try:
                if len(match) == 3:
                    year = int(match[0]) if len(match[0]) == 4 else date.today().year
                    month = int(match[1])
                    day = int(match[2])
                    dates.append(date(year, month, day))
                elif len(match) == 2:
                    month = int(match[0])
                    day = int(match[1])
                    dates.append(date(date.today().year, month, day))
            except ValueError:
                continue
    
    return dates


@app.post("/save-language-config")
def save_language_config_endpoint() -> Response:
    try:
        project_key = request.form["project"]
        language_names = request.form.getlist("language_names[]")
        language_prices = request.form.getlist("language_prices[]")
        default_languages = request.form.getlist("default_languages[]")
        
        languages = {}
        for name, price in zip(language_names, language_prices):
            if name.strip():
                languages[name.strip()] = float(price)
        
        save_language_config(project_key, languages, default_languages)
        return {"status": "success"}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


@app.post("/reset-language-config")
def reset_language_config_endpoint() -> Response:
    try:
        project_key = request.form["project"]
        reset_to_default(project_key)
        return {"status": "success"}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


@app.post("/save-save-path")
def save_save_path_endpoint() -> Response:
    try:
        project_key = request.form["project_key"]
        save_path = request.form["save_path"]

        set_save_path(project_key, save_path)
        return {"status": "success"}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


@app.get("/api/base-paths")
def get_base_paths() -> Response:
    try:
        return jsonify({
            "status": "success",
            "quote_history_base": str(get_quote_history_dir()),
            "settlement_base": str(get_settlement_dir()),
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/base-paths")
def save_base_paths() -> Response:
    try:
        import json as _json
        config_path = ROOT / "config" / "base_paths.json"
        data = request.get_json() or {}
        config = {}
        if config_path.exists():
            try:
                config = _json.loads(config_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        if "quote_history_base" in data:
            config["quote_history_base"] = data["quote_history_base"]
        if "settlement_base" in data:
            config["settlement_base"] = data["settlement_base"]
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(_json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        return jsonify({"status": "success"})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/save-project-config")
def save_project_config_endpoint() -> Response:
    try:
        if request.is_json:
            config_data = request.get_json()
            for item in config_data:
                save_project_config(
                    item["project_key"],
                    item["display_name"],
                    item["sort_order"],
                    item.get("company", ""),
                )
        else:
            project_key = request.form["project_key"]
            display_name = request.form["display_name"]
            sort_order = int(request.form["sort_order"])
            company = request.form.get("company", "")
            save_project_config(project_key, display_name, sort_order, company)
        
        return {"status": "success"}
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


@app.get("/report")
def report_page() -> str:
    """报价报告页：查看所有项目的报价单记录"""
    projects_by_company = get_projects_by_company()
    all_projects = []
    for company, projects in projects_by_company.items():
        all_projects.extend(projects)
    selected = all_projects[0] if all_projects else None
    return render_template("report.html", projects_by_company=projects_by_company, selected_project=selected)


@app.get("/api/all-quotes")
def all_quotes() -> Response:
    """返回所有报价单的平铺列表，按修改时间倒序"""
    history_dir = get_quote_history_dir()
    result = []
    if not history_dir.exists():
        return {"status": "success", "data": result}
    for xlsx_file in sorted(history_dir.rglob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
        rel = xlsx_file.relative_to(history_dir)
        result.append({
            "name": xlsx_file.name,
            "path": str(xlsx_file),
            "size": xlsx_file.stat().st_size,
            "mtime": xlsx_file.stat().st_mtime,
            "project": rel.parts[0] if len(rel.parts) > 1 else "",
        })
    return {"status": "success", "data": result}


@app.get("/api/auto-quote-logs")
def auto_quote_logs() -> Response:
    """返回自动报价执行记录"""
    import json as _json
    index_file = OUTPUTS_DIR / "logs" / "auto_quote" / "_index.json"
    if not index_file.exists():
        return {"status": "success", "data": []}
    try:
        index = _json.loads(index_file.read_text(encoding="utf-8"))
    except Exception:
        index = []
    # 倒序（最新的在前）
    index.reverse()
    return {"status": "success", "data": index}


@app.get("/api/auto-quote-log-content")
def auto_quote_log_content() -> Response:
    """返回某条执行记录的完整日志内容"""
    log_path = request.args.get("path", "")
    if not log_path:
        return {"status": "error", "message": "缺少 path 参数"}, 400
    p = Path(log_path)
    if not p.exists():
        return {"status": "error", "message": "日志文件不存在"}, 404
    # 安全检查：只允许在 logs 目录下
    try:
        p.relative_to(OUTPUTS_DIR / "logs")
    except ValueError:
        return {"status": "error", "message": "路径不允许"}, 403
    return {"status": "success", "content": p.read_text(encoding="utf-8")}


@app.get("/api/quote-records")
def quote_records_api() -> Response:
    from settlement.settlement_tracker import load_records, TRACKER_DIR
    project_key = request.args.get("project", "")
    if project_key:
        records = load_records(project_key)
    else:
        records = []
        if TRACKER_DIR.exists():
            for d in TRACKER_DIR.iterdir():
                if d.is_dir() and (d / "records.json").exists():
                    records.extend(load_records(d.name))
    records.sort(key=lambda r: r.get("quote_date", ""), reverse=True)
    return {"status": "success", "data": records}


@app.get("/download")
def download() -> Response:
    path = resolve_allowed_path(request.args["path"])
    return send_file(path, as_attachment=True, download_name=path.name)


def _save_auto_quote_log(project_key: str, output: str, status: str):
    """保存自动报价执行日志"""
    log_dir = OUTPUTS_DIR / "logs" / "auto_quote"
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"{project_key}_{ts}.log"
    log_file.write_text(output, encoding="utf-8")
    # 同时维护一个摘要索引
    index_file = log_dir / "_index.json"
    import json as _json
    if index_file.exists():
        try:
            index = _json.loads(index_file.read_text(encoding="utf-8"))
        except Exception:
            index = []
    else:
        index = []
    index.append({
        "project": project_key,
        "timestamp": datetime.now().isoformat(),
        "status": status,
        "log_file": str(log_file),
        "summary": output.strip().split("\n")[-1] if output.strip() else "",
    })
    # 只保留最近 500 条
    if len(index) > 500:
        index = index[-500:]
    index_file.write_text(_json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")


@app.post("/auto-quote")
def auto_quote() -> Response:
    """完美世界自动报价：从飞书表格读取需求 → 下载HTML → 生成报价单 → 上传回飞书"""
    import io
    import sys
    project_key = request.form.get("project", "huanta")
    try:
        old_stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()
        auto_quote_run(project_key)
        output = buffer.getvalue()
        _save_auto_quote_log(project_key, output, "success")
        return jsonify({"status": "success", "message": output})
    except Exception as exc:
        import traceback
        err = traceback.format_exc()
        _save_auto_quote_log(project_key, err, "error")
        return jsonify({"status": "error", "message": str(exc)}), 500
    finally:
        sys.stdout = old_stdout


@app.post("/zhan-shuang/auto-quote")
def zhan_shuang_auto_quote() -> Response:
    """战双邮件自动报价"""
    import io
    import sys
    try:
        old_stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()
        zhan_shuang_auto_quote_run()
        output = buffer.getvalue()
        _save_auto_quote_log("zhan_shuang", output, "success")
        return jsonify({"status": "success", "message": output})
    except Exception as exc:
        import traceback
        err = traceback.format_exc()
        _save_auto_quote_log("zhan_shuang", err, "error")
        return jsonify({"status": "error", "message": str(exc)}), 500
    finally:
        sys.stdout = old_stdout


@app.post("/zhan-shuang/feishu-quote")
def zhan_shuang_feishu_quote() -> Response:
    """战双飞书双月汇总报价"""
    import io
    import sys
    try:
        old_stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()
        zhan_shuang_feishu_run()
        output = buffer.getvalue()
        _save_auto_quote_log("zhan_shuang_feishu", output, "success")
        return jsonify({"status": "success", "message": output})
    except Exception as exc:
        import traceback
        err = traceback.format_exc()
        _save_auto_quote_log("zhan_shuang_feishu", err, "error")
        return jsonify({"status": "error", "message": str(exc)}), 500
    finally:
        sys.stdout = old_stdout


# ======================== 叠纸批次报价 ========================

@app.get("/diezhi/batch-info")
def diezhi_batch_info() -> Response:
    """获取叠纸项目的批次信息"""
    project_key = request.args.get("project", "")
    if not project_key:
        return jsonify({"status": "error", "message": "缺少 project 参数"}), 400
    try:
        sys.modules.pop('quote_system.auto_quote_diezhi', None)
        from quote_system.auto_quote_diezhi import detect_batches
        info = detect_batches(project_key)
        return jsonify({"status": "success", **info})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/diezhi/generate-quote")
def diezhi_generate_quote() -> Response:
    """生成叠纸项目的批次报价单"""
    import io
    import sys
    try:
        data = request.get_json() or {}
        project_key = data.get("project", "")
        batch_no = data.get("batch_no")
        quote_date_str = data.get("quote_date")
        price = data.get("price")
        service_type = data.get("service_type", "初翻")

        if not project_key or not batch_no:
            return jsonify({"status": "error", "message": "缺少 project 或 batch_no 参数"}), 400

        quote_date = parse_date(quote_date_str) if quote_date_str else date.today()
        batch_no = int(batch_no)

        old_stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()
        try:
            out_path = generate_batch_quote(
                project_key=project_key,
                batch_no=batch_no,
                quote_date=quote_date,
                price=float(price) if price is not None else 0.32,
                service_type=service_type,
            )
            output = buffer.getvalue()

            try:
                from settlement.settlement_tracker import quick_record
                proj = resolve_project(project_key)
                _add_quote_record(quick_record(
                    project_key=project_key,
                    company=proj.company or "叠纸",
                    req_name=out_path.stem,
                    word_count=0,
                    total_price=0,
                    quote_file=out_path.name,
                    language=service_type,
                    source="batch",
                ))
            except Exception:
                pass

            return jsonify({
                "status": "success",
                "output_path": str(out_path),
                "output_name": out_path.name,
                "message": output,
            })
        finally:
            sys.stdout = old_stdout
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/tk-fill/start")
def start_tk_fill() -> Response:
    """启动 TK 自动填表：API 方式下载 N 列 HTML → 解析 → 填写 O-Y。"""
    global TK_FILL_PROCESS
    try:
        if TK_FILL_PROCESS and TK_FILL_PROCESS.poll() is None:
            return jsonify({
                "status": "error",
                "message": "TK 填表已经在运行，请等待完成。",
            }), 409

        args = [
            sys.executable,
            "-m",
            "quote_system.auto_fill_tk",
        ]
        tail_text = request.form.get("tail", "").strip()
        if tail_text:
            try:
                tail = int(tail_text)
                if tail > 0:
                    args.extend(["--tail", str(tail)])
            except ValueError:
                pass
        if request.form.get("dry_run") == "1":
            args.append("--dry-run")

        log_dir = OUTPUTS_DIR / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / "tk_fill.log"
        log_file = log_path.open("a", encoding="utf-8")
        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        TK_FILL_PROCESS = subprocess.Popen(
            args,
            cwd=str(ROOT),
            stdout=log_file,
            stderr=subprocess.STDOUT,
            creationflags=creationflags,
        )
        return jsonify({
            "status": "success",
            "message": "已启动 TK 自动填表。正在下载 HTML → 解析 → 填写 O-Y 列。",
            "log_path": str(log_path),
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/fill-4399/start")
def start_fill_4399() -> Response:
    """启动 4399 填表：下载 G 列 HTML → 解析 → 填写 K/L/M 列。"""
    global FILL_4399_PROCESS
    try:
        if FILL_4399_PROCESS and FILL_4399_PROCESS.poll() is None:
            return jsonify({
                "status": "error",
                "message": "4399 填表已经在运行，请等待完成。",
            }), 409

        args = [
            sys.executable,
            "-m",
            "quote_system.auto_quote_4399",
        ]
        if request.form.get("dry_run") == "1":
            args.append("--dry-run")

        log_dir = OUTPUTS_DIR / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / "4399_fill.log"
        log_file = log_path.open("a", encoding="utf-8")
        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        FILL_4399_PROCESS = subprocess.Popen(
            args,
            cwd=str(ROOT),
            stdout=log_file,
            stderr=subprocess.STDOUT,
            creationflags=creationflags,
        )
        return jsonify({
            "status": "success",
            "message": "已启动 4399 自动填表。正在下载 HTML → 解析 → 填写 K/L/M 列。",
            "log_path": str(log_path),
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/quote-history")
def quote_history() -> Response:
    try:
        project_key = request.args.get("project", "")
        history_dir = get_quote_history_dir()
        result = []
        
        if not history_dir.exists():
            return {"status": "success", "data": result, "project": project_key}
        
        if project_key:
            try:
                project_config = resolve_project(project_key)
                target_dir = history_dir / project_config.history_dir
            except ValueError:
                target_dir = history_dir / project_key
            
            if target_dir.exists() and target_dir.is_dir():
                result = build_history_tree(target_dir)
        else:
            for project_dir in sorted(history_dir.iterdir(), key=lambda x: x.name):
                if not project_dir.is_dir():
                    continue

                project_info = {
                    "name": project_dir.name,
                    "path": str(project_dir),
                    "folders": []
                }

                # 根目录下的 xlsx 文件
                root_files = sorted(
                    [f for f in project_dir.glob("*.xlsx") if f.is_file()],
                    key=lambda x: x.stat().st_mtime, reverse=True
                )
                if root_files:
                    root_folder = {"name": "根目录", "path": str(project_dir), "files": []}
                    for f in root_files:
                        root_folder["files"].append({
                            "name": f.name, "path": str(f),
                            "size": f.stat().st_size, "mtime": f.stat().st_mtime
                        })
                    project_info["folders"].append(root_folder)

                # 子目录
                for sub_dir in sorted(
                    [d for d in project_dir.iterdir() if d.is_dir()],
                    key=lambda x: x.name, reverse=True
                ):
                    sub_info = {"name": sub_dir.name, "path": str(sub_dir), "files": []}
                    for f in sorted(
                        [f for f in sub_dir.rglob("*.xlsx") if f.is_file()],
                        key=lambda x: x.stat().st_mtime, reverse=True
                    ):
                        sub_info["files"].append({
                            "name": f.name, "path": str(f),
                            "size": f.stat().st_size, "mtime": f.stat().st_mtime
                        })
                    if sub_info["files"]:
                        project_info["folders"].append(sub_info)

                if project_info["folders"]:
                    result.append(project_info)
        
        return {"status": "success", "data": result, "project": project_key}
    except Exception as exc:
        print(f"获取报价单历史错误: {exc}")
        return {"status": "error", "message": str(exc)}, 500


def build_history_tree(target_dir):
    result = [{
        "name": target_dir.name,
        "path": str(target_dir),
        "folders": []
    }]

    # 根目录下的 xlsx 文件
    root_files = sorted(
        [f for f in target_dir.glob("*.xlsx") if f.is_file()],
        key=lambda x: x.stat().st_mtime, reverse=True
    )
    if root_files:
        root_info = {
            "name": "根目录",
            "path": str(target_dir),
            "files": []
        }
        for file in root_files:
            root_info["files"].append({
                "name": file.name,
                "path": str(file),
                "size": file.stat().st_size,
                "mtime": file.stat().st_mtime
            })
        result[0]["folders"].append(root_info)

    # 检查子目录（如"已结算"）
    sub_dirs = sorted(
        [d for d in target_dir.iterdir() if d.is_dir()],
        key=lambda x: x.name, reverse=True
    )
    for sub_dir in sub_dirs:
        sub_info = {
            "name": sub_dir.name,
            "path": str(sub_dir),
            "files": []
        }
        xlsx_files = sorted(
            [f for f in sub_dir.rglob("*.xlsx") if f.is_file()],
            key=lambda x: x.stat().st_mtime, reverse=True
        )
        for file in xlsx_files:
            sub_info["files"].append({
                "name": file.name,
                "path": str(file),
                "size": file.stat().st_size,
                "mtime": file.stat().st_mtime
            })
        if sub_info["files"]:
            result[0]["folders"].append(sub_info)

    return result


@app.post("/delete-quote")
def delete_quote() -> Response:
    try:
        data = request.get_json() or {}
        file_path_str = data.get("path", "")

        if not file_path_str:
            return jsonify({"status": "error", "message": "路径不能为空"}), 400

        file_path = resolve_allowed_path(file_path_str)

        if not file_path.is_file():
            return jsonify({"status": "error", "message": "只能删除文件"}), 400
        if file_path.suffix.lower() not in DELETABLE_SUFFIXES:
            return jsonify({"status": "error", "message": "不允许删除该类型文件"}), 400

        file_path.unlink()
        print(f"已删除文件: {file_path}")

        return jsonify({"status": "success"})
    except Exception as exc:
        print(f"删除报价单错误: {exc}")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/open-folder")
def open_folder():
    import subprocess
    import platform

    raw_path = request.args.get("path", "")

    if not raw_path:
        return jsonify({"status": "error", "message": "路径不能为空"}), 400

    try:
        file_path_obj = resolve_allowed_path(raw_path)
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    try:
        system = platform.system()
        if system == "Windows":
            if file_path_obj.is_dir():
                subprocess.Popen(["explorer", str(file_path_obj)])
            else:
                subprocess.Popen(["explorer", f"/select,{file_path_obj}"])
        elif system == "Darwin":
            if file_path_obj.is_dir():
                subprocess.Popen(["open", str(file_path_obj)])
            else:
                subprocess.Popen(["open", "-R", str(file_path_obj)])
        else:
            target = str(file_path_obj) if file_path_obj.is_dir() else str(file_path_obj.parent)
            subprocess.Popen(["xdg-open", target])

        return jsonify({"status": "success"})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


def _find_wps() -> str | None:
    """查找 WPS 可执行文件路径"""
    import os as _os
    local_appdata = _os.environ.get("LOCALAPPDATA", "")
    candidates = []
    if local_appdata:
        ks = Path(local_appdata) / "Kingsoft" / "WPS Office"
        if ks.exists():
            for d in sorted(ks.iterdir(), reverse=True):
                wps = d / "office6" / "wps.exe"
                if wps.exists():
                    candidates.append(str(wps))
    for base in (_os.environ.get("PROGRAMFILES(X86)", ""), _os.environ.get("PROGRAMFILES", ""),
                 r"C:\Program Files (x86)", r"C:\Program Files"):
        if not base:
            continue
        ks = Path(base) / "Kingsoft" / "WPS Office"
        if ks.exists():
            for d in sorted(ks.iterdir(), reverse=True):
                wps = d / "office6" / "wps.exe"
                if wps.exists():
                    candidates.append(str(wps))
    return candidates[0] if candidates else None


@app.route("/open-file", methods=["GET", "POST"])
def open_file():
    """直接用 WPS 打开文件（Excel/Word等），找不到 WPS 则用默认程序"""
    import platform

    if request.method == "POST":
        data = request.get_json() or {}
        raw_path = data.get("path", "")
    else:
        raw_path = request.args.get("path", "")

    if not raw_path:
        return jsonify({"status": "error", "message": "路径不能为空"}), 400

    try:
        file_path_obj = resolve_allowed_path(raw_path)
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    if file_path_obj.is_dir():
        return jsonify({"status": "error", "message": "路径是目录"}), 400

    try:
        system = platform.system()
        if system == "Windows":
            wps = _find_wps()
            if wps:
                subprocess.Popen([wps, str(file_path_obj)])
            else:
                os.startfile(str(file_path_obj))
        elif system == "Darwin":
            subprocess.Popen(["open", str(file_path_obj)])
        else:
            subprocess.Popen(["xdg-open", str(file_path_obj)])

        return jsonify({"status": "success"})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500




# ======================== 结算模块 ========================

# ---- Bilibili 项目账单（马娘/邦邦2/HBR） ----

@app.get("/api/settlement/mamian/preview")
def mamian_bill_preview() -> Response:
    """获取 Bilibili 项目指定月份的账单预览"""
    from settlement.generate_settlement_mamian import scan_quotes, _get_bill_config
    try:
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        project_key = request.args.get("project", "umamusume")
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = scan_quotes(year, month, project_key)
        total_words = sum(r["word_count"] for r in records)
        total_amount = sum(r["total_price"] for r in records)
        cfg = _get_bill_config(project_key)

        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "total_words": total_words,
            "total_amount": round(total_amount, 2),
            "project": project_key,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement/mamian/generate-bill")
def mamian_generate_bill() -> Response:
    """生成 Bilibili 项目月度账单"""
    try:
        from settlement.generate_settlement_mamian import (
            scan_quotes, MamianSettlementGenerator, _get_bill_config, move_settled,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        project_key = data.get("project", "umamusume")
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = scan_quotes(year, month, project_key)
        if not records:
            cfg = _get_bill_config(project_key)
            return jsonify({"status": "error", "message": f"{year}年{month}月没有未结算的 {cfg['project_full_name']} 交付记录"}), 404

        gen = MamianSettlementGenerator(year, month, project_key)
        output_path = gen.generate(records, None)

        move_settled(records, year, month, project_key)

        # 开票金额 = 所有报价单合计行 K 列的总和
        total_amount = sum(r["total_price"] for r in records)
        invoice_text = _generate_invoice_text(total_amount)
        _save_invoice_request(year, month, project_key, invoice_text)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的账单",
            "count": len(records),
            "file": {"path": str(output_path), "name": output_path.name},
            "total_words": sum(r["word_count"] for r in records),
            "total_amount": total_amount,
            "invoice_text": invoice_text,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement/mamian/generate-sealed")
def mamian_generate_sealed() -> Response:
    """生成 Bilibili 项目盖章版结算单（对账单确认后调用，总金额自动从对账单读取）"""
    try:
        from settlement.generate_settlement_sealed_mamian import generate_sealed, _read_total_from_bill, _get_sealed_config
        from settlement.generate_settlement_mamian import _get_bill_config
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        project_key = data.get("project", "umamusume")
        total_amount = data.get("total_amount")
        if total_amount is not None:
            total_amount = float(total_amount)
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        output_path = generate_sealed(year, month, total_amount, project_key=project_key)

        # 同步更新开票请求
        bill_cfg = _get_bill_config(project_key)
        sealed_cfg = _get_sealed_config(project_key)
        output_dir = get_settlement_dir() / f"{year}年{month}月" / "Bilibili" / bill_cfg["project"]
        sealed_amount = _read_total_from_bill(output_dir, year, month, sealed_cfg["code"])
        invoice_text = _generate_invoice_text(sealed_amount)
        _save_invoice_request(year, month, project_key, invoice_text)

        return jsonify({
            "status": "success",
            "message": f"盖章版结算单已生成并打开",
            "file": {"path": str(output_path), "name": output_path.name},
            "invoice_text": invoice_text,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


def _check_discrepancies(project_key: str, settle_records: list[dict], year: int, month: int) -> list[dict]:
    try:
        from settlement.settlement_tracker import get_unsettled
        quote_recs = get_unsettled(project_key, year, month)
    except Exception:
        return []
    discrepancies = []
    for sr in settle_records:
        sr_name = (sr.get("req_name") or sr.get("task_no") or "").strip()
        sr_words = sr.get("billable_words") or sr.get("word_count") or 0
        for qr in quote_recs:
            qr_name = (qr.get("req_name") or "").strip()
            if not sr_name or not qr_name:
                continue
            if sr_name in qr_name or qr_name in sr_name:
                qr_words = qr.get("billable_words") or qr.get("word_count") or 0
                if qr_words > 0 and sr_words > 0:
                    diff_pct = abs(sr_words - qr_words) / qr_words * 100
                    if diff_pct > 5:
                        discrepancies.append({
                            "req_name": sr_name,
                            "settle_words": sr_words,
                            "quote_words": qr_words,
                            "diff_pct": round(diff_pct, 1),
                        })
                break
    return discrepancies


@app.get("/api/settlement/preview")
def settlement_preview() -> Response:
    """获取指定月份的结算预览数据"""
    try:
        from settlement.generate_settlement import read_feishu_data
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        discrepancies = _check_discrepancies("huanta", records, year, month)
        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "discrepancies": discrepancies,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/api/settlement_yh_games/preview")
def settlement_yh_games_preview() -> Response:
    """获取异环游戏内指定月份的结算预览数据"""
    try:
        from settlement.generate_settlement_yh_games import read_feishu_data
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        discrepancies = _check_discrepancies("yihuan_nei", records, year, month)
        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "discrepancies": discrepancies,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        error_msg = str(exc)
        if "错误页面" in error_msg or "权限" in error_msg:
            return jsonify({
                "status": "error",
                "message": f"飞书表格访问失败，可能是表格权限问题或Sheet ID不正确。错误：{error_msg}",
                "is_permission_error": True
            }), 500
        return jsonify({"status": "error", "message": error_msg}), 500


@app.get("/api/settlement_yh_publish/preview")
def settlement_yh_publish_preview() -> Response:
    """获取异环发行指定月份的结算预览数据"""
    try:
        from settlement.generate_settlement_yh_publish import read_feishu_data
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        discrepancies = _check_discrepancies("yihuan_faxing", records, year, month)
        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "discrepancies": discrepancies,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        error_msg = str(exc)
        if "错误页面" in error_msg or "权限" in error_msg:
            return jsonify({
                "status": "error",
                "message": f"飞书表格访问失败，可能是表格权限问题或Sheet ID不正确。错误：{error_msg}",
                "is_permission_error": True
            }), 500
        return jsonify({"status": "error", "message": error_msg}), 500


@app.post("/api/settlement/generate")
def settlement_generate() -> Response:
    """生成结算单Excel和验收单Word"""
    try:
        from settlement.generate_settlement import (
            read_feishu_data, generate_settlement_excel, generate_acceptance_docx,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        if not records:
            return jsonify({"status": "error", "message": f"{year}年{month}月没有交付记录"}), 404

        output_dir = get_settlement_dir() / f"{year}年{month}月" / "完美世界" / "幻塔"
        output_dir.mkdir(parents=True, exist_ok=True)

        excel_path = generate_settlement_excel(records, year, month, output_dir)
        docx_path = generate_acceptance_docx(records, year, month, output_dir)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的结算文件",
            "count": len(records),
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "docx": {"path": str(docx_path), "name": docx_path.name},
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        import sys
        sys.stderr.flush()
        return jsonify({"status": "error", "message": f"{type(exc).__name__}: {str(exc)}"}), 500


@app.post("/api/settlement_yh_games/generate")
def settlement_yh_games_generate() -> Response:
    """生成异环游戏内结算单Excel和验收单Word"""
    try:
        from settlement.generate_settlement_yh_games import (
            read_feishu_data as read_yh_games_data,
            generate_settlement_excel as generate_yh_games_excel,
            generate_acceptance_docx as generate_yh_games_docx,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_yh_games_data(year, month)
        if not records:
            return jsonify({"status": "error", "message": f"{year}年{month}月没有交付记录"}), 404

        output_dir = get_settlement_dir() / f"{year}年{month}月" / "完美世界" / "异环游戏内"
        output_dir.mkdir(parents=True, exist_ok=True)

        excel_path = generate_yh_games_excel(records, year, month, output_dir)
        docx_path = generate_yh_games_docx(records, year, month, output_dir)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的异环游戏内结算文件",
            "count": len(records),
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "docx": {"path": str(docx_path), "name": docx_path.name},
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        import sys
        sys.stderr.flush()
        return jsonify({"status": "error", "message": f"{type(exc).__name__}: {str(exc)}"}), 500


@app.post("/api/settlement_yh_publish/generate")
def settlement_yh_publish_generate() -> Response:
    """生成异环发行结算单Excel和验收单Word"""
    try:
        from settlement.generate_settlement_yh_publish import (
            read_feishu_data as read_yh_publish_data,
            generate_settlement_excel as generate_yh_publish_excel,
            generate_acceptance_docx as generate_yh_publish_docx,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_yh_publish_data(year, month)
        if not records:
            return jsonify({"status": "error", "message": f"{year}年{month}月没有交付记录"}), 404

        output_dir = get_settlement_dir() / f"{year}年{month}月" / "完美世界" / "异环发行"
        output_dir.mkdir(parents=True, exist_ok=True)

        excel_path = generate_yh_publish_excel(records, year, month, output_dir)
        docx_path = generate_yh_publish_docx(records, year, month, output_dir)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的异环发行结算文件",
            "count": len(records),
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "docx": {"path": str(docx_path), "name": docx_path.name},
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        import sys
        sys.stderr.flush()
        return jsonify({"status": "error", "message": f"{type(exc).__name__}: {str(exc)}"}), 500


@app.post("/api/settlement/batch-perfect-world")
def settlement_batch_perfect_world() -> Response:
    """一键生成完美世界三个项目的结算文件"""
    try:
        from settlement.generate_settlement import (
            read_feishu_data as read_huanta,
            generate_settlement_excel as excel_huanta,
            generate_acceptance_docx as docx_huanta,
        )
        from settlement.generate_settlement_yh_games import (
            read_feishu_data as read_yh_games,
            generate_settlement_excel as excel_yh_games,
            generate_acceptance_docx as docx_yh_games,
        )
        from settlement.generate_settlement_yh_publish import (
            read_feishu_data as read_yh_publish,
            generate_settlement_excel as excel_yh_publish,
            generate_acceptance_docx as docx_yh_publish,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        results = {}
        base_dir = get_settlement_dir() / f"{year}年{month}月" / "完美世界"

        _MODULE_SHEET = {
            "幻塔": ("generate_settlement", "ZXBogz"),
            "异环游戏内": ("generate_settlement_yh_games", "bfba7c"),
            "异环发行": ("generate_settlement_yh_publish", "S5yHmP"),
        }
        for name, reader, excel_fn, docx_fn, subdir in [
            ("幻塔", read_huanta, excel_huanta, docx_huanta, "幻塔"),
            ("异环游戏内", read_yh_games, excel_yh_games, docx_yh_games, "异环游戏内"),
            ("异环发行", read_yh_publish, excel_yh_publish, docx_yh_publish, "异环发行"),
        ]:
            try:
                records = reader(year, month)
                if not records:
                    results[name] = {"status": "empty", "message": f"{name}: {year}年{month}月没有交付记录"}
                    continue
                output_dir = base_dir / subdir
                output_dir.mkdir(parents=True, exist_ok=True)
                excel_path = excel_fn(records, year, month, output_dir)
                docx_path = docx_fn(records, year, month, output_dir)
                _mark_pw_settled(records, _MODULE_SHEET[name][1])
                results[name] = {
                    "status": "success",
                    "count": len(records),
                    "excel": {"path": str(excel_path), "name": excel_path.name},
                    "docx": {"path": str(docx_path), "name": docx_path.name},
                }
            except Exception as e:
                results[name] = {"status": "error", "message": f"{name}: {str(e)}"}

        return jsonify({"status": "success", "results": results})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/api/settlement/pw-notification-text")
def settlement_pw_notification_text() -> Response:
    try:
        from settlement.generate_settlement import read_feishu_data as read_huanta
        from settlement.generate_settlement_yh_games import read_feishu_data as read_yh_games
        from settlement.generate_settlement_yh_publish import read_feishu_data as read_yh_publish

        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        texts = []
        for name, reader, tpl in [
            ("异环游戏内", read_yh_games, "yihuan_nei"),
            ("幻塔", read_huanta, "huanta"),
            ("异环发行", read_yh_publish, "yihuan_faxing"),
        ]:
            records = reader(year, month)
            if not records:
                continue
            total_amount = sum(float(r.get("amount") or 0) for r in records)
            total_words = sum(float(r.get("billable_words") or 0) for r in records)

            if tpl == "yihuan_nei":
                text = (
                    f"@Juvenile  @Castiel @吴睿立 @陈宏昭\n"
                    f"各位老师好。\n"
                    f"异环项目游戏内日语翻译和校对需求的{month}月份账单和验收单准备好了。\n\n"
                    f"有劳确认。\n"
                    f"---------------------------------------------------------------------------\n"
                    f"{total_amount:.2f} 元 - NTE - {month}月日语翻译/校对需求\n"
                    f"字数：{total_words:.1f}字\n"
                    f"---------------------------------------------------------------------------"
                )
            elif tpl == "huanta":
                text = (
                    f"@Juvenile  @Castiel @吴睿立 @陈宏昭\n"
                    f"各位老师好。\n"
                    f"幻塔项目的游戏内日语翻译和校对需求的{month}月份账单和验收单准备好了。\n\n"
                    f"有劳确认。\n"
                    f"----------------------------------------------------------------\n"
                    f"{total_amount:.2f}元 - TOF - {month}月日语翻译/校对需求\n"
                    f"字数：{total_words:.1f}字\n"
                    f"----------------------------------------------------------------"
                )
            else:
                text = (
                    f"陈老师，\n"
                    f"您好。异环游戏外日语翻译需求的{month}月份的账单和验收单做好了。有劳确认。\n"
                    f"---------------------------------------------------------------------\n"
                    f"{total_amount:.2f}元 - NTE-发行-{month}月日语翻译/校对需求\n"
                    f"字数：{total_words:.1f}字\n"
                    f"---------------------------------------------------------------------"
                )
            texts.append({"project": name, "text": text, "amount": round(total_amount, 2), "words": round(total_words, 1)})

        return jsonify({"status": "success", "texts": texts})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ---- 战双版更结算 ----

@app.get("/api/settlement_zhan_shuang/preview")
def settlement_zhan_shuang_preview() -> Response:
    try:
        from settlement.generate_settlement_zhan_shuang import read_feishu_data
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        total = sum(r['word_count'] for r in records)
        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "total_words": total,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement_zhan_shuang/generate")
def settlement_zhan_shuang_generate() -> Response:
    try:
        import sys
        sys.modules.pop('settlement.generate_settlement_zhan_shuang', None)
        from settlement.generate_settlement_zhan_shuang import (
            read_feishu_data, generate_settlement_excel,
        )
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, month)
        if not records:
            return jsonify({"status": "error", "message": f"{year}年{month}月没有战双版更交付记录"}), 404

        output_dir = get_settlement_dir() / f"{year}年{month}月" / "库洛游戏" / "战双版更"
        output_dir.mkdir(parents=True, exist_ok=True)

        excel_path = generate_settlement_excel(records, year, month, output_dir)

        total_amount = sum(r['word_count'] * 0.64 for r in records)
        invoice_text = _generate_kuro_invoice_text(total_amount)
        _save_kuro_invoice_request(year, month, "战双版更", invoice_text)

        quote_dir = get_quote_history_dir() / "库洛游戏" / "战双版更"
        _move_settled_quotes(quote_dir, [r['file_name'] for r in records if r.get('file_name')], year, month)
        _update_zs_status(records)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的战双版更结算单",
            "count": len(records),
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "total_amount": round(total_amount, 2),
            "total_words": sum(r['word_count'] for r in records),
            "invoice_text": invoice_text,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ---- 战双发行结算 ----

@app.get("/api/settlement_zhan_shuang_faxing/preview")
def settlement_zhan_shuang_faxing_preview() -> Response:
    try:
        from settlement.generate_settlement_zhan_shuang_faxing import read_feishu_data
        year = int(request.args.get("year", 0))
        months_str = request.args.get("months", "")
        months = [int(m.strip()) for m in months_str.split(",") if m.strip()]
        if not (2020 <= year <= 2099 and months and all(1 <= m <= 12 for m in months)):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, months)
        total_words = sum(r['word_count'] for r in records)
        total_amount_pretax = sum(r['word_count'] * r['unit_price'] for r in records)
        total_amount = round(total_amount_pretax * 1.06, 2)
        return jsonify({
            "status": "success",
            "data": records,
            "count": len(records),
            "total_words": total_words,
            "total_amount": total_amount,
            "total_amount_pretax": round(total_amount_pretax, 2),
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement_zhan_shuang_faxing/generate")
def settlement_zhan_shuang_faxing_generate() -> Response:
    try:
        import sys as _sys
        _sys.modules.pop('settlement.generate_settlement_zhan_shuang_faxing', None)
        from settlement.generate_settlement_zhan_shuang_faxing import (
            read_feishu_data, generate_settlement_excel, _month_label,
        )
        from quote_system.auto_quote_zhan_shuang_feishu import get_bi_monthly_label_for_date
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        months = data.get("months", [])
        if not isinstance(months, list):
            months = [int(m.strip()) for m in str(months).split(",") if m.strip()]
        months = [int(m) for m in months]
        if not (2020 <= year <= 2099 and months and all(1 <= m <= 12 for m in months)):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        records = read_feishu_data(year, months)
        if not records:
            month_label = _month_label(months)
            return jsonify({"status": "error", "message": f"{year}年{month_label}月没有战双发行交付记录"}), 404

        month_label = _month_label(months)

        # 找到对应的报价单文件
        quote_dir = get_quote_history_dir() / "库洛游戏" / "战双发行"
        target_quote = None
        for f in quote_dir.glob("*.xlsx"):
            if "已结算" in str(f):
                continue
            # 检查文件名是否匹配结算月份
            for m in months:
                if f"{year}年{m}月" in f.name or f"{year}年{months[0]}＆{months[1]}月" in f.name:
                    target_quote = f
                    break
            if target_quote:
                break

        # 如果没有找到精确匹配，尝试根据月份标签找
        if not target_quote:
            for f in quote_dir.glob("*.xlsx"):
                if "已结算" in str(f):
                    continue
                if month_label in f.name:
                    target_quote = f
                    break

        # 检查报价单中是否有未结算的需求
        unsettled_names = []
        if target_quote:
            quote_names = _get_quote_names_in_file(target_quote)
            settled_names = {r['file_name'] for r in records if r.get('file_name')}
            unsettled_names = list(quote_names - settled_names)

        if unsettled_names:
            return jsonify({
                "status": "warning",
                "message": f"报价单中还有 {len(unsettled_names)} 条未结算的需求，无法移动报价单",
                "unsettled": unsettled_names[:10],  # 最多显示10条
                "count": len(unsettled_names),
            }), 409

        output_dir = get_settlement_dir() / f"{year}年{months[-1]}月" / "库洛游戏" / "战双发行"
        output_dir.mkdir(parents=True, exist_ok=True)

        excel_path = generate_settlement_excel(records, year, months, output_dir)

        total_amount = sum(r['word_count'] * r['unit_price'] for r in records)
        total_words = sum(r['word_count'] for r in records)
        invoice_text = _generate_kuro_invoice_text(total_amount)
        _save_kuro_invoice_request(year, months[-1], "战双发行", invoice_text)

        # 移动整个报价单到已结算目录
        if target_quote:
            settled_dir = quote_dir / "已结算" / f"{year}年{months[-1]}月"
            settled_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(str(target_quote), str(settled_dir / target_quote.name))
            print(f"已移动报价单到: {settled_dir / target_quote.name}")

        _update_zs_status(records)

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(records)} 条记录的战双发行结算单",
            "count": len(records),
            "excel": {"path": str(excel_path), "name": excel_path.name},
            "total_amount": round(total_amount, 2),
            "total_words": total_words,
            "invoice_text": invoice_text,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ---- 4399 结算 ----


@app.get("/api/settlement_4399/preview")
def settlement_4399_preview() -> Response:
    try:
        from settlement.generate_settlement_4399 import (
            read_feishu_data_4399, read_feishu_data_boqi, _get_project_code,
        )
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        main_data = read_feishu_data_4399(year, month)
        boqi_records = read_feishu_data_boqi(year, month)
        if boqi_records:
            main_data["波奇"] = boqi_records

        projects_preview = {}
        for proj_name, records in main_data.items():
            if not records:
                continue
            code = _get_project_code(proj_name)
            total_words = sum(r["word_count"] for r in records)
            total_cny = total_words * 0.52
            projects_preview[proj_name] = {
                "code": code,
                "count": len(records),
                "total_words": total_words,
                "total_cny": round(total_cny, 2),
                "records": [{
                    "req_name": r["req_name"],
                    "deliv_date": (excel_serial_to_date(r["deliv_date"]) or datetime(2000, 1, 1)).strftime("%Y-%m-%d"),
                    "word_count": r["word_count"],
                } for r in records],
            }

        total_usd = 0
        from settlement.generate_settlement_4399 import get_exchange_rate
        from datetime import datetime as _dt
        auto_rate = get_exchange_rate(_dt.now().year, _dt.now().month)
        return jsonify({
            "status": "success",
            "data": projects_preview,
            "project_count": len(projects_preview),
            "exchange_rate": auto_rate,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


def excel_serial_to_date(serial) -> datetime | None:
    from datetime import timedelta
    try:
        num = float(serial)
    except (TypeError, ValueError):
        return None
    base = datetime(1899, 12, 30)
    try:
        return base + timedelta(days=num)
    except Exception:
        return None


def _mark_4399_settled(year, month):
    """4399表O列写"已请款"。"""
    from .feishu_client import FeishuClient
    from settlement.generate_settlement_4399 import (
        read_feishu_data_4399,
        SPREADSHEET_TOKEN_4399, SHEET_ID_4399,
    )

    client_4399 = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN_4399, sheet_id=SHEET_ID_4399)
    main_data = read_feishu_data_4399(year, month)
    for records in main_data.values():
        for r in records:
            if 'row_index' not in r:
                continue
            try:
                client_4399.write_cell(r['row_index'], 14, "已请款")
            except Exception as e:
                print(f"4399 行{r['row_index']} 标记已请款失败: {e}")


def _verify_4399_settled(year, month):
    """验证4399项目O列是否已标记为已请款，返回失败的行号列表。"""
    from .feishu_client import FeishuClient
    from settlement.generate_settlement_4399 import (
        read_feishu_data_4399,
        SPREADSHEET_TOKEN_4399, SHEET_ID_4399,
    )

    client_4399 = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN_4399, sheet_id=SHEET_ID_4399)
    main_data = read_feishu_data_4399(year, month)
    failed = []
    for records in main_data.values():
        for r in records:
            if 'row_index' not in r:
                continue
            try:
                val = client_4399.read_cell(r['row_index'], 14, SHEET_ID_4399)
                if str(val).strip() != "已请款":
                    failed.append(r['row_index'])
            except Exception:
                failed.append(r['row_index'])
    return failed


@app.post("/api/settlement_4399/generate")
def settlement_4399_generate() -> Response:
    try:
        import sys as _sys
        _sys.modules.pop('settlement.generate_settlement_4399', None)
        from settlement.generate_settlement_4399 import generate_all
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        exchange_rate = float(data.get("exchange_rate", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400
        if exchange_rate <= 0:
            from settlement.generate_settlement_4399 import get_exchange_rate
            from datetime import datetime as _dt
            exchange_rate = get_exchange_rate(_dt.now().year, _dt.now().month) or 0
        if exchange_rate <= 0:
            return jsonify({"status": "error", "message": "请输入有效汇率"}), 400

        result = generate_all(year, month, exchange_rate)
        _mark_4399_settled(year, month)

        files = []
        table1 = []
        table2 = []
        for proj_name, info in result.items():
            files.append({
                "project": proj_name,
                "xlsx": {"path": info["xlsx"], "name": Path(info["xlsx"]).name},
                "settlement_pdf": {"path": info["settlement_pdf"], "name": Path(info["settlement_pdf"]).name},
                "invoice_pdf": {"path": info["invoice_pdf"], "name": Path(info["invoice_pdf"]).name},
                "usd_amount": info["usd_amount"],
            })
            if proj_name in ("指尖无双", "主宰世界"):
                table1.append({"name": proj_name, "usd_amount": info["usd_amount"]})
            else:
                table2.append({"name": proj_name, "usd_amount": info["usd_amount"]})

        return jsonify({
            "status": "success",
            "message": f"已生成 {len(files)} 个项目的结算文件",
            "files": files,
            "table1": table1,
            "table2": table2,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ---- 叠纸结算（恋与深空/闪暖/ニキ新作）统一入口 ----

@app.get("/api/settlement_diezhi/preview")
def settlement_diezhi_preview() -> Response:
    try:
        import sys as _sys
        _sys.modules.pop('settlement.generate_settlement_diezhi', None)
        from settlement.generate_settlement_diezhi import get_all_preview
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        preview = get_all_preview(year, month)
        total_amount = round(sum(p['total_amount'] for p in preview.values()), 2)
        return jsonify({
            "status": "success",
            "data": preview,
            "total_amount": total_amount,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement_diezhi/generate")
def settlement_diezhi_generate() -> Response:
    try:
        import sys as _sys
        _sys.modules.pop('settlement.generate_settlement_diezhi', None)
        from settlement.generate_settlement_diezhi import generate_all, get_invoice_info
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        results = generate_all(year, month)
        invoice_text = get_invoice_info(year, month)

        files = []
        for key, info in results.items():
            if info['path']:
                files.append({
                    'project': key,
                    'path': info['path'],
                    'name': Path(info['path']).name,
                    'amount': info['amount'],
                })

        total_amount = round(sum(info['amount'] for info in results.values()), 2)
        return jsonify({
            "status": "success",
            "message": f"已生成 {len(files)} 个项目结算单",
            "files": files,
            "total_amount": total_amount,
            "invoice": invoice_text,
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/api/settlement_diezhi/invoice-info")
def settlement_diezhi_invoice_info() -> Response:
    try:
        from settlement.generate_settlement_diezhi import get_invoice_info
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        text = get_invoice_info(year, month)
        return jsonify({
            "status": "success",
            "invoice": text or "",
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ======================== TK 结算 ========================

@app.post("/api/settlement/tk/copy-orderlist")
def tk_copy_orderlist() -> Response:
    """复制 TK 订单明细表模板到结算文件夹"""
    try:
        from settlement.generate_settlement_tk import copy_orderlist
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        output_path = copy_orderlist(year, month)
        os.startfile(str(output_path))
        return jsonify({
            "status": "success",
            "message": f"已复制订单明细表模板",
            "file": {"path": str(output_path), "name": output_path.name},
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement/tk/generate")
def tk_generate_settlement() -> Response:
    """生成 TK 结算单 PDF + Invoice PDF"""
    try:
        from settlement.generate_settlement_tk import generate_settlement_pdf, generate_invoice_pdf
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        settlement_path = generate_settlement_pdf(year, month)
        invoice_path = generate_invoice_pdf(year, month)

        return jsonify({
            "status": "success",
            "message": f"已生成结算单和Invoice",
            "settlement": {"path": str(settlement_path), "name": settlement_path.name},
            "invoice": {"path": str(invoice_path), "name": invoice_path.name},
        })
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement/zulong/generate")
def settlement_zulong_generate() -> Response:
    try:
        from settlement.generate_settlement_zulong import generate_zulong_settlement
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        counts = data.get("counts", {})
        yishan = int(counts.get("yishan", 0) or 0)
        longzu = int(counts.get("longzu", 0) or 0)
        if yishan <= 0 and longzu <= 0:
            return jsonify({"status": "error", "message": "至少需要填写一个项目的件数"}), 400

        files = generate_zulong_settlement(year, month, {"yishan": yishan, "longzu": longzu})
        return jsonify({"status": "success", "files": files})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ======================== 结算文件列表 ========================

@app.get("/api/settlement-files")
def settlement_files() -> Response:
    """列出结算目录中的文件，按项目/年月筛选"""
    project_key = request.args.get("project", "")
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    if not project_key:
        return jsonify({"status": "error", "message": "缺少 project 参数"}), 400

    settlement_dir = get_settlement_dir()
    result: list[dict] = []

    bill_project_dirs = {
        "umamusume": "马娘",
        "bang2": "bang2",
        "hbr": "HBR",
    }
    perfect_world_dirs = {
        "huanta": "幻塔",
        "yihuan_nei": "异环游戏内",
        "yihuan_faxing": "异环发行",
    }
    kuluo_dirs = {
        "zhan_shuang": "战双版更",
        "zhan_shuang_faxing": "战双发行",
    }
    diezhi_dirs = {
        "liandishenkong": "叠纸",
        "shining_nikki": "叠纸",
        "niki_xinzuo": "叠纸",
    }
    month_display = month.replace(",", "&")
    if project_key in bill_project_dirs:
        search_dir = settlement_dir / f"{year}年{month_display}月" / "Bilibili" / bill_project_dirs[project_key]
    elif project_key in perfect_world_dirs:
        search_dir = settlement_dir / f"{year}年{month_display}月" / "完美世界" / perfect_world_dirs[project_key]
    elif project_key in kuluo_dirs:
        search_dir = settlement_dir / f"{year}年{month_display}月" / "库洛游戏" / kuluo_dirs[project_key]
    elif project_key in diezhi_dirs:
        search_dir = settlement_dir / f"{year}年{month_display}月" / diezhi_dirs[project_key]
    elif project_key == "tk":
        search_dir = settlement_dir / f"{year}年{month_display}月" / "TK"
    elif project_key == "4399":
        search_dir = settlement_dir / f"{year}年{month_display}月" / "4399"
    elif project_key == "zulong":
        search_dir = settlement_dir / f"{year}年{month_display}月" / "祖龙"
    else:
        search_dir = settlement_dir

    if search_dir.exists():
        for f in sorted(search_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if not f.is_file():
                continue
            if project_key not in bill_project_dirs and project_key not in perfect_world_dirs and project_key not in kuluo_dirs and project_key != "tk" and year and month:
                ym = f"{year}年{month_display}月" if month else str(year)
                if ym not in f.name and f"{year}" not in f.name:
                    continue
            result.append({
                "name": f.name,
                "path": str(f),
                "size": f.stat().st_size,
                "mtime": f.stat().st_mtime,
            })

    return jsonify({"status": "success", "data": result, "project": project_key})


# ======================== 开票请求 ========================

INVOICE_COMPANY = "上海哔哩哔哩科技有限公司"
INVOICE_TAX_ID = "91310000MA1G8B8AXD"
INVOICE_ITEM = "翻译费"
INVOICE_ADDRESS = "上海市杨浦区四平路1945号1641室"
INVOICE_PHONE = "021-25099255"
INVOICE_BANK = "招商银行股份有限公司上海自贸试验区分行"
INVOICE_ACCOUNT = "121923952410103 121923952410208"

KURO_INVOICE_COMPANY = "广州库洛科技有限公司"
KURO_INVOICE_TAX_ID = "91440101MA59JPNR3D"
KURO_INVOICE_ITEM = "现代服务*翻译服务费"
KURO_INVOICE_ADDRESS = "广州市天河区棠下荷光三横路7号之一107房"
KURO_INVOICE_PHONE = "020-85655091"
KURO_INVOICE_BANK_ACCOUNT = "120912143410201"
KURO_INVOICE_BANK = "招行广州科技园支行"


def _generate_invoice_text(total_amount: float) -> str:
    return (
        f"麻烦开1张电子专票\n"
        f"贵司名称：{INVOICE_COMPANY}\n"
        f"纳税人识别号：{INVOICE_TAX_ID}\n"
        f"开票项目：{INVOICE_ITEM}\n"
        f"地址：{INVOICE_ADDRESS}\n"
        f"电话：{INVOICE_PHONE}\n"
        f"银行账号：{INVOICE_ACCOUNT}\n"
        f"开户银行：{INVOICE_BANK}\n"
        f"金额：{total_amount:.2f}"
    )


def _save_invoice_request(year: int, month: int, project_key: str, text: str) -> Path:
    from settlement.generate_settlement_mamian import _get_bill_config
    cfg = _get_bill_config(project_key)
    inv_dir = get_settlement_dir() / f"{year}年{month}月" / "Bilibili" / cfg["project"]
    inv_dir.mkdir(parents=True, exist_ok=True)
    inv_path = inv_dir / "invoice_request.txt"
    inv_path.write_text(text, encoding="utf-8")
    return inv_path


def _load_invoice_request(year: int, month: int, project_key: str) -> str | None:
    from settlement.generate_settlement_mamian import _get_bill_config
    cfg = _get_bill_config(project_key)
    inv_path = get_settlement_dir() / f"{year}年{month}月" / "Bilibili" / cfg["project"] / "invoice_request.txt"
    if inv_path.exists():
        return inv_path.read_text(encoding="utf-8")
    return None


def _generate_kuro_invoice_text(total_amount: float) -> str:
    return (
        f"贵司名称：{KURO_INVOICE_COMPANY}\n"
        f"纳税人识别号：{KURO_INVOICE_TAX_ID}\n"
        f"开票项目：{KURO_INVOICE_ITEM}\n"
        f"地址：{KURO_INVOICE_ADDRESS}\n"
        f"电话：{KURO_INVOICE_PHONE}\n"
        f"银行账号：{KURO_INVOICE_BANK_ACCOUNT}\n"
        f"开户银行：{KURO_INVOICE_BANK}\n"
        f"金额：{total_amount:.2f}"
    )


def _save_kuro_invoice_request(year: int, month: int, project_name: str, text: str) -> Path:
    inv_dir = get_settlement_dir() / f"{year}年{month}月" / "库洛游戏" / project_name
    inv_dir.mkdir(parents=True, exist_ok=True)
    inv_path = inv_dir / "invoice_request.txt"
    inv_path.write_text(text, encoding="utf-8")
    return inv_path


def _load_kuro_invoice_request(year: int, month: int, project_name: str) -> str | None:
    inv_path = get_settlement_dir() / f"{year}年{month}月" / "库洛游戏" / project_name / "invoice_request.txt"
    if inv_path.exists():
        return inv_path.read_text(encoding="utf-8")
    return None


ZS_SPREADSHEET_TOKEN = "Wup0wnUPIiiIr2k8T23c4zASnjd"
ZS_SHEET_ID = "WmpiFq"


def _update_zs_status(records):
    """将战双在线表 F 列从"已报价"改为"已结算"。"""
    from .feishu_client import FeishuClient
    client = FeishuClient(spreadsheet_token=ZS_SPREADSHEET_TOKEN, sheet_id=ZS_SHEET_ID)
    for r in records:
        if 'row_index' not in r:
            continue
        client.write_cell(r['row_index'], 5, "已结算")


def _move_settled_quotes(quote_dir, file_names, year, month):
    if not file_names:
        return
    src_dir = Path(quote_dir)
    settled_dir = src_dir / "已结算" / f"{year}年{month}月"
    settled_dir.mkdir(parents=True, exist_ok=True)
    for fname in file_names:
        src = src_dir / fname
        if src.exists():
            shutil.move(str(src), str(settled_dir / fname))


def _get_quote_names_in_file(xlsx_path: Path) -> set[str]:
    """读取报价单中的所有需求名称"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    names = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v:
            names.add(str(v).strip())
    wb.close()
    return names


@app.get("/api/invoice-request")
def get_invoice_request() -> Response:
    """获取已保存的开票请求"""
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    project_key = request.args.get("project", "umamusume")
    if not year or not month:
        return jsonify({"status": "error", "message": "缺少年月参数"}), 400

    if project_key == "zhan_shuang":
        text = _load_kuro_invoice_request(int(year), int(month), "战双版更")
    elif project_key == "zhan_shuang_faxing":
        text = _load_kuro_invoice_request(int(year), int(month), "战双发行")
    else:
        text = _load_invoice_request(int(year), int(month), project_key)

    if text:
        return jsonify({"status": "success", "text": text})
    return jsonify({"status": "success", "text": None})


@app.post("/api/invoice-request")
def generate_invoice_request() -> Response:
    """从对账单读取金额，生成开票请求"""
    try:
        from settlement.generate_settlement_sealed_mamian import _read_total_from_bill, _get_sealed_config
        from settlement.generate_settlement_mamian import _get_bill_config
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        project_key = data.get("project", "umamusume")
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        bill_cfg = _get_bill_config(project_key)
        sealed_cfg = _get_sealed_config(project_key)
        output_dir = get_settlement_dir() / f"{year}年{month}月" / "Bilibili" / bill_cfg["project"]
        total_amount = _read_total_from_bill(output_dir, year, month, sealed_cfg["code"])
        text = _generate_invoice_text(total_amount)
        _save_invoice_request(year, month, project_key, text)

        return jsonify({"status": "success", "text": text, "total_amount": total_amount})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ======================== 一键报价 ========================

_quote_jobs: dict = {}
_quote_jobs_lock = threading.Lock()

AUTO_QUOTE_PROJECTS = [
    {"key": "huanta", "name": "幻塔", "company": "完美世界"},
    {"key": "yihuan_nei", "name": "异环游戏内", "company": "完美世界"},
    {"key": "yihuan_faxing", "name": "异环发行", "company": "完美世界"},
    {"key": "zhan_shuang", "name": "战双版更", "company": "库洛游戏"},
    {"key": "zhan_shuang_feishu", "name": "战双发行", "company": "库洛游戏"},
    {"key": "tk", "name": "TK填表", "company": "Bilibili"},
    {"key": "4399", "name": "4399填表", "company": "4399"},
]


def _execute_single_auto_quote(project_key: str):
    if project_key in ("huanta", "yihuan_nei", "yihuan_faxing"):
        auto_quote_run(project_key)
    elif project_key == "zhan_shuang":
        zhan_shuang_auto_quote_run()
    elif project_key == "zhan_shuang_feishu":
        zhan_shuang_feishu_run()
    elif project_key == "tk":
        from .auto_fill_tk import run_scheduled
        run_scheduled()
    elif project_key == "4399":
        from .auto_quote_4399 import run_scheduled
        run_scheduled()
    else:
        raise ValueError(f"项目 {project_key} 不支持自动报价")


@app.post("/api/quote-all")
def start_quote_all() -> Response:
    data = request.get_json() or {}
    selected = data.get("projects", [])
    if not selected:
        return jsonify({"status": "error", "message": "请选择至少一个项目"}), 400

    job_id = uuid.uuid4().hex[:12]
    valid_keys = {p["key"] for p in AUTO_QUOTE_PROJECTS}
    projects_status = {}
    for key in selected:
        if key not in valid_keys:
            continue
        proj = next(p for p in AUTO_QUOTE_PROJECTS if p["key"] == key)
        projects_status[key] = {"name": proj["name"], "company": proj["company"], "status": "pending", "output": "", "error": None}

    with _quote_jobs_lock:
        _quote_jobs[job_id] = {"status": "running", "projects": projects_status}

    def _run_all():
        job = _quote_jobs[job_id]
        for key in selected:
            if key not in valid_keys:
                continue
            with _quote_jobs_lock:
                job["projects"][key]["status"] = "running"
            try:
                buf = io.StringIO()
                old_stdout = sys.stdout
                sys.stdout = buf
                try:
                    _execute_single_auto_quote(key)
                    with _quote_jobs_lock:
                        job["projects"][key]["output"] = buf.getvalue()
                        job["projects"][key]["status"] = "done"
                finally:
                    sys.stdout = old_stdout
            except Exception as e:
                import traceback
                with _quote_jobs_lock:
                    job["projects"][key]["status"] = "error"
                    job["projects"][key]["error"] = str(e)
                    job["projects"][key]["output"] = traceback.format_exc()
        with _quote_jobs_lock:
            job["status"] = "done"

    thread = threading.Thread(target=_run_all, daemon=True)
    thread.start()

    return jsonify({"status": "success", "job_id": job_id, "projects": list(projects_status.values())})


@app.get("/api/quote-all/status")
def quote_all_status() -> Response:
    job_id = request.args.get("job_id", "")
    with _quote_jobs_lock:
        if not job_id or job_id not in _quote_jobs:
            return jsonify({"status": "error", "message": "无效的 job_id"}), 400
        job = dict(_quote_jobs[job_id])
    return jsonify({"status": "success", "job": job})


# ======================== 月度结算工作台 ========================

SETTLEMENT_WORKBENCH_PROJECTS = [
    {"key": "huanta", "name": "幻塔", "company": "完美世界", "type": "perfect_world", "module": "generate_settlement", "subdir": "幻塔"},
    {"key": "yihuan_nei", "name": "异环游戏内", "company": "完美世界", "type": "perfect_world", "module": "generate_settlement_yh_games", "subdir": "异环游戏内"},
    {"key": "yihuan_faxing", "name": "异环发行", "company": "完美世界", "type": "perfect_world", "module": "generate_settlement_yh_publish", "subdir": "异环发行"},
    {"key": "zhan_shuang", "name": "战双版更", "company": "库洛游戏", "type": "zhan_shuang", "module": "generate_settlement_zhan_shuang", "subdir": "战双版更"},
    {"key": "zhan_shuang_faxing", "name": "战双发行", "company": "库洛游戏", "type": "zhan_shuang_faxing", "module": "generate_settlement_zhan_shuang_faxing", "subdir": "战双发行"},
    {"key": "umamusume", "name": "马娘", "company": "Bilibili", "type": "mamian", "project_key": "umamusume"},
    {"key": "bang2", "name": "BANG2", "company": "Bilibili", "type": "mamian", "project_key": "bang2"},
    {"key": "hbr", "name": "炽焰天穹", "company": "Bilibili", "type": "mamian", "project_key": "hbr"},
    {"key": "liandishenkong", "name": "恋与深空", "company": "叠纸", "type": "diezhi"},
    {"key": "shining_nikki", "name": "闪暖", "company": "叠纸", "type": "diezhi"},
    {"key": "niki_xinzuo", "name": "ニキ新作", "company": "叠纸", "type": "diezhi"},
    {"key": "tk", "name": "TK项目", "company": "Bilibili", "type": "tk"},
    {"key": "4399", "name": "4399", "company": "4399", "type": "4399"},
]


def _settle_preview_pw(module_name, year, month):
    import importlib
    mod = importlib.import_module(f"settlement.{module_name}")
    records = mod.read_feishu_data(year, month)
    count = len(records)
    total_words = sum(r.get("billable_words", 0) or 0 for r in records)
    total_amount = sum(r.get("amount", 0) or 0 for r in records)
    return {"count": count, "total_words": total_words, "total_amount": round(total_amount, 2)}


def _mark_pw_settled(records, sheet_id):
    """完美世界项目J列写"已请款"。"""
    from .feishu_client import FeishuClient
    client = FeishuClient(sheet_id=sheet_id)
    for r in records:
        if 'row_index' not in r:
            continue
        try:
            client.write_cell(r['row_index'], 9, "已请款")
        except Exception as e:
            print(f"完美世界 行{r['row_index']} 标记已请款失败: {e}")


def _verify_pw_settled(records, sheet_id):
    """验证完美世界项目J列是否已标记为已请款，返回失败的行号列表。"""
    from .feishu_client import FeishuClient
    client = FeishuClient(sheet_id=sheet_id)
    failed = []
    for r in records:
        if 'row_index' not in r:
            continue
        try:
            val = client.read_cell(r['row_index'], 9, sheet_id)
            if str(val).strip() != "已请款":
                failed.append(r['row_index'])
        except Exception:
            failed.append(r['row_index'])
    return failed


def _check_settlement_exists(output_dir, prefix):
    """检查当月结算文件是否已存在"""
    if not output_dir.exists():
        return False
    for f in output_dir.iterdir():
        if f.is_file() and f.name.startswith(prefix) and not f.name.startswith("~$"):
            return True
    return False


_PW_SHEET_IDS = {
    'generate_settlement': 'ZXBogz',
    'generate_settlement_yh_games': 'bfba7c',
    'generate_settlement_yh_publish': 'S5yHmP',
}


def _settle_generate_pw(module_name, year, month, subdir, company):
    import importlib
    mod = importlib.import_module(f"settlement.{module_name}")
    records = mod.read_feishu_data(year, month)
    if not records:
        return {"status": "empty", "message": "没有交付记录"}
    output_dir = get_settlement_dir() / f"{year}年{month}月" / company / subdir
    output_dir.mkdir(parents=True, exist_ok=True)
    if _check_settlement_exists(output_dir, "【"):
        return {"status": "exists", "message": f"{subdir}当月已有结算文件，请检查是否重复结算"}
    excel_path = mod.generate_settlement_excel(records, year, month, output_dir)
    docx_path = mod.generate_acceptance_docx(records, year, month, output_dir)
    sheet_id = _PW_SHEET_IDS.get(module_name)
    verification = {"feishu_mark": "skipped"}
    if sheet_id:
        _mark_pw_settled(records, sheet_id)
        failed_rows = _verify_pw_settled(records, sheet_id)
        verification["feishu_mark"] = "ok" if not failed_rows else f"failed: {failed_rows}"
    total_amount = sum(r.get("amount", 0) or 0 for r in records)
    total_words = sum(r.get("billable_words", 0) or 0 for r in records)
    return {
        "status": "ok", "count": len(records),
        "total_amount": round(total_amount, 2), "total_words": total_words,
        "verification": verification,
        "files": [
            {"name": excel_path.name, "path": str(excel_path)},
            {"name": docx_path.name, "path": str(docx_path)},
        ],
    }


def _settle_preview_zs(module_name, year, month, multi_month):
    import importlib
    mod = importlib.import_module(f"settlement.{module_name}")
    if multi_month:
        records = mod.read_feishu_data(year, [month])
        total_amount_pretax = sum(r.get("word_count", 0) * r.get("unit_price", 0) for r in records)
    else:
        records = mod.read_feishu_data(year, month)
        total_amount_pretax = sum(r.get("word_count", 0) * 0.64 for r in records)
    total_amount = round(total_amount_pretax * 1.06, 2)
    return {"count": len(records), "total_words": sum(r.get("word_count", 0) for r in records), "total_amount": total_amount, "total_amount_pretax": round(total_amount_pretax, 2)}


def _settle_generate_zs(module_name, year, month, subdir, multi_month):
    import importlib
    mod = importlib.import_module(f"settlement.{module_name}")
    if multi_month:
        records = mod.read_feishu_data(year, [month])
        total_amount = sum(r.get("word_count", 0) * r.get("unit_price", 0) for r in records)
    else:
        records = mod.read_feishu_data(year, month)
        total_amount = sum(r.get("word_count", 0) * 0.64 for r in records)
    if not records:
        return {"status": "empty", "message": "没有交付记录"}
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "库洛游戏" / subdir
    output_dir.mkdir(parents=True, exist_ok=True)
    if _check_settlement_exists(output_dir, "【"):
        return {"status": "exists", "message": f"{subdir}当月已有结算文件，请检查是否重复结算"}
    if multi_month:
        excel_path = mod.generate_settlement_excel(records, year, [month], output_dir)
    else:
        excel_path = mod.generate_settlement_excel(records, year, month, output_dir)
    return {
        "status": "ok", "count": len(records),
        "total_amount": round(total_amount, 2), "total_words": sum(r.get("word_count", 0) for r in records),
        "files": [{"name": excel_path.name, "path": str(excel_path)}],
    }


def _settle_preview_mamian(year, month, project_key):
    from settlement.generate_settlement_mamian import scan_quotes
    records = scan_quotes(year, month, project_key)
    total_words = sum(r.get("word_count", 0) for r in records)
    total_amount = sum(r.get("total_price", 0.0) for r in records)
    return {"count": len(records), "total_words": total_words, "total_amount": round(total_amount, 2)}


def _settle_generate_mamian(year, month, project_key):
    from settlement.generate_settlement_mamian import scan_quotes, MamianSettlementGenerator, BILL_CONFIG
    cfg = BILL_CONFIG.get(project_key, {})
    project_name = cfg.get("project", project_key)
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "Bilibili" / project_name
    output_dir.mkdir(parents=True, exist_ok=True)
    if _check_settlement_exists(output_dir, "大连"):
        return {"status": "exists", "message": f"{project_name}当月已有结算文件，请检查是否重复结算"}
    records = scan_quotes(year, month, project_key)
    if not records:
        return {"status": "empty", "message": "没有找到未结算记录"}
    gen = MamianSettlementGenerator(year, month, project_key)
    import argparse
    args = argparse.Namespace(year=year, month=month, project=project_key, dry_run=False)
    path = gen.generate(records, args)
    total_amount = sum(r.get("total_price", 0.0) for r in records)
    total_words = sum(r.get("word_count", 0) for r in records)
    return {
        "status": "ok", "count": len(records),
        "total_amount": round(total_amount, 2), "total_words": total_words,
        "files": [{"name": path.name, "path": str(path)}],
    }


def _settle_preview_4399(year, month):
    from settlement.generate_settlement_4399 import read_feishu_data_4399, read_feishu_data_boqi
    main_data = read_feishu_data_4399(year, month)
    boqi_data = read_feishu_data_boqi(year, month)
    total_count = 0
    total_words = 0
    for records in main_data.values():
        total_count += len(records)
        total_words += sum(r.get("word_count", 0) or 0 for r in records)
    if boqi_data:
        total_count += len(boqi_data)
        total_words += sum(r.get("word_count", 0) or 0 for r in boqi_data)
    total_amount = total_words * 0.52
    return {"count": total_count, "total_words": total_words, "total_amount": round(total_amount, 2)}


def _settle_generate_4399(year, month, exchange_rate):
    from settlement.generate_settlement_4399 import generate_all
    from pathlib import Path as _Path
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "4399"
    output_dir.mkdir(parents=True, exist_ok=True)
    if _check_settlement_exists(output_dir, "【"):
        return {"status": "exists", "message": "4399当月已有结算文件，请检查是否重复结算"}
    result = generate_all(year, month, exchange_rate)
    _mark_4399_settled(year, month)
    failed_rows = _verify_4399_settled(year, month)
    verification = {"feishu_mark": "ok" if not failed_rows else f"failed: {failed_rows}"}
    files = []
    total_usd = 0.0
    table1 = []
    table2 = []
    for proj_name, info in result.items():
        files.append({"project": proj_name, "name": _Path(info["settlement_pdf"]).name, "path": info["settlement_pdf"]})
        total_usd += float(info.get("usd_amount", 0))
        if proj_name in ("指尖无双", "主宰世界"):
            table1.append({"name": proj_name, "usd_amount": info["usd_amount"]})
        else:
            table2.append({"name": proj_name, "usd_amount": info["usd_amount"]})
    return {"status": "ok", "files": files, "total_amount": round(total_usd, 2), "table1": table1, "table2": table2, "verification": verification}


def _settle_generate_tk(year, month):
    from settlement.generate_settlement_tk import generate_settlement_pdf, generate_invoice_pdf
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "TK"
    output_dir.mkdir(parents=True, exist_ok=True)
    if _check_settlement_exists(output_dir, "【"):
        return {"status": "exists", "message": "TK当月已有结算文件，请检查是否重复结算"}
    settlement_path = generate_settlement_pdf(year, month)
    invoice_path = generate_invoice_pdf(year, month)
    return {
        "status": "ok",
        "files": [
            {"name": settlement_path.name, "path": str(settlement_path)},
            {"name": invoice_path.name, "path": str(invoice_path)},
        ],
    }


_diezhi_preview_result: dict = {}
_diezhi_generate_result: dict = {}


def _settle_preview_diezhi(year, month):
    from settlement.generate_settlement_diezhi import get_all_preview
    preview = get_all_preview(year, month)
    result = {}
    for key, info in preview.items():
        result[key] = {
            "count": len(info.get("quotes", [])),
            "total_words": info.get("total_words", 0),
            "total_amount": info.get("total_amount", 0),
        }
    return result


def _settle_generate_diezhi(year, month):
    from settlement.generate_settlement_diezhi import generate_all
    from pathlib import Path as _Path
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "叠纸"
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.is_file() and f.name.startswith("【") and not f.name.startswith("~$"):
                return {"status": "exists", "message": "叠纸当月已有结算文件，请检查是否重复结算"}
    results = generate_all(year, month)
    files = []
    total_amount = 0.0
    for key, info in results.items():
        total_amount += info.get("amount", 0)
        if info.get("path"):
            files.append({"project": key, "name": _Path(info["path"]).name, "path": info["path"]})
    return {"status": "ok", "files": files, "total_amount": round(total_amount, 2)}


@app.get("/api/settlement-workbench/preview")
def settlement_workbench_preview() -> Response:
    try:
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        results = []
        diezhi_done = False
        diezhi_data = {}

        for proj in SETTLEMENT_WORKBENCH_PROJECTS:
            try:
                if proj["type"] == "diezhi":
                    if not diezhi_done:
                        diezhi_data = _settle_preview_diezhi(year, month)
                        diezhi_done = True
                    info = diezhi_data.get(proj["key"], {"count": 0, "total_words": 0, "total_amount": 0})
                    results.append({**proj, **info, "status": "has_data" if info.get("count", 0) > 0 else "empty"})
                elif proj["type"] == "perfect_world":
                    info = _settle_preview_pw(proj["module"], year, month)
                    results.append({**proj, **info, "status": "has_data" if info["count"] > 0 else "empty"})
                elif proj["type"] in ("zhan_shuang", "zhan_shuang_faxing"):
                    multi = proj["type"] == "zhan_shuang_faxing"
                    info = _settle_preview_zs(proj["module"], year, month, multi)
                    results.append({**proj, **info, "status": "has_data" if info["count"] > 0 else "empty"})
                elif proj["type"] == "mamian":
                    info = _settle_preview_mamian(year, month, proj["project_key"])
                    results.append({**proj, **info, "status": "has_data" if info["count"] > 0 else "empty"})
                elif proj["type"] == "4399":
                    info = _settle_preview_4399(year, month)
                    results.append({**proj, **info, "status": "has_data" if info["count"] > 0 else "empty"})
                elif proj["type"] == "tk":
                    results.append({**proj, "count": 0, "total_words": 0, "total_amount": 0, "status": "manual"})
            except Exception as e:
                results.append({**proj, "count": 0, "total_words": 0, "total_amount": 0, "status": "error", "error": str(e)})

        return jsonify({"status": "success", "projects": results, "year": year, "month": month})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.post("/api/settlement-workbench/generate")
def settlement_workbench_generate() -> Response:
    try:
        data = request.get_json() or {}
        year = int(data.get("year", 0))
        month = int(data.get("month", 0))
        selected = data.get("projects", [])
        exchange_rate = data.get("exchange_rate")
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400
        if not selected:
            return jsonify({"status": "error", "message": "请选择至少一个项目"}), 400

        results = {}
        diezhi_done = False

        for key in selected:
            proj = next((p for p in SETTLEMENT_WORKBENCH_PROJECTS if p["key"] == key), None)
            if not proj:
                results[key] = {"status": "error", "message": f"未知项目: {key}"}
                continue
            try:
                if proj["type"] == "diezhi":
                    if diezhi_done:
                        continue
                    diezhi_done = True
                    result = _settle_generate_diezhi(year, month)
                    for dk in ["liandishenkong", "shining_nikki", "niki_xinzuo"]:
                        if dk in selected:
                            results[dk] = result
                elif proj["type"] == "perfect_world":
                    results[key] = _settle_generate_pw(proj["module"], year, month, proj["subdir"], proj["company"])
                elif proj["type"] == "zhan_shuang":
                    results[key] = _settle_generate_zs(proj["module"], year, month, proj["subdir"], False)
                elif proj["type"] == "zhan_shuang_faxing":
                    results[key] = _settle_generate_zs(proj["module"], year, month, proj["subdir"], True)
                elif proj["type"] == "mamian":
                    results[key] = _settle_generate_mamian(year, month, proj["project_key"])
                elif proj["type"] == "4399":
                    rate = float(exchange_rate) if exchange_rate else 0
                    results[key] = _settle_generate_4399(year, month, rate)
                elif proj["type"] == "tk":
                    results[key] = _settle_generate_tk(year, month)
            except Exception as e:
                import traceback
                results[key] = {"status": "error", "message": f"{type(e).__name__}: {str(e)}"}

        return jsonify({"status": "success", "results": results, "year": year, "month": month})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ======================== 结算文件树 ========================


@app.get("/api/settlement-workbench/files")
def settlement_workbench_files() -> Response:
    try:
        year = int(request.args.get("year", 0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"status": "error", "message": "年月参数无效"}), 400

        base = get_settlement_dir() / f"{year}年{month}月"
        if not base.exists():
            return jsonify({"status": "success", "tree": [], "year": year, "month": month})

        tree = []
        for company_dir in sorted(base.iterdir()):
            if not company_dir.is_dir():
                continue
            company_node = {"name": company_dir.name, "type": "dir", "children": []}
            has_subdir = False
            for project_dir in sorted(company_dir.iterdir()):
                if not project_dir.is_dir():
                    continue
                has_subdir = True
                project_node = {"name": project_dir.name, "type": "dir", "children": []}
                for f in sorted(project_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
                    if not f.is_file():
                        continue
                    project_node["children"].append({
                        "name": f.name, "type": "file", "path": str(f),
                        "size": f.stat().st_size, "mtime": f.stat().st_mtime,
                    })
                if project_node["children"]:
                    company_node["children"].append(project_node)
            if not has_subdir:
                for f in sorted(company_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
                    if not f.is_file():
                        continue
                    company_node["children"].append({
                        "name": f.name, "type": "file", "path": str(f),
                        "size": f.stat().st_size, "mtime": f.stat().st_mtime,
                    })
            if company_node["children"]:
                tree.append(company_node)

        return jsonify({"status": "success", "tree": tree, "year": year, "month": month})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(exc)}), 500


# ======================== 工具函数 ========================

def save_uploads(files) -> list[Path]:
    batch_dir = UPLOAD_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S") / uuid.uuid4().hex[:8]
    batch_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    for file in files:
        if not file or not file.filename:
            continue
        suffix = Path(file.filename).suffix.lower()
        if suffix not in {".html", ".htm"}:
            raise ValueError(f"只支持 MemoQ HTML 文件：{file.filename}")
        target = batch_dir / safe_upload_name(file.filename)
        file.save(target)
        saved.append(target)
    if not saved:
        shutil.rmtree(batch_dir, ignore_errors=True)
    return saved


def safe_upload_name(filename: str) -> str:
    cleaned = "".join(ch if ch not in '<>:"/\\|?*' else "_" for ch in filename.strip())
    return cleaned or f"memoq_{uuid.uuid4().hex}.html"


def parse_language_prices(project_key: str) -> tuple[list[str] | None, dict[str, float]]:
    languages = [value.strip() for value in request.form.getlist("selected_languages") if value.strip()]
    prices: dict[str, float] = {}
    for language, price_text in zip(
        request.form.getlist("selected_languages"),
        request.form.getlist("selected_prices"),
    ):
        language = language.strip()
        price_text = price_text.strip()
        if not language or not price_text:
            continue
        try:
            prices[language] = float(price_text)
        except ValueError as exc:
            raise ValueError(f"{language} 的单价格式不正确: {price_text}") from exc
    if project_key == "zhan_shuang":
        return None, prices
    return (languages or None), prices


def summarize_stats(project_key: str, result) -> str:
    if not result.stats:
        return ""
    if project_key == "zhan_shuang":
        total = 0
        for stats in result.stats:
            for row in stats.rows:
                label = row.type.strip().lower()
                if label in {"85%-94%", "75%-84%", "50%-74%", "no match"}:
                    total += row.source_asian_characters or 0
        return f"计费字数约 {total}。"
    try:
        words = sum(quote_words(stats) for stats in result.stats)
        return f"报价字数约 {words}。"
    except Exception:
        return f"已解析 {len(result.stats)} 个 HTML。"


def resolve_allowed_path(raw_path: str) -> Path:
    path = Path(raw_path).resolve()
    allowed_roots = [(ROOT / "outputs").resolve(), (get_quote_history_dir()).resolve(), (get_settlement_dir()).resolve()]
    _add_save_path_roots(allowed_roots)
    if not any(is_relative_to(path, allowed_root) for allowed_root in allowed_roots):
        raise ValueError("不允许访问该路径")
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")
    return path


def _add_save_path_roots(allowed_roots: list[Path]):
    try:
        from .save_path_config import CONFIG_PATH
        if CONFIG_PATH.exists():
            import json
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                configs = json.load(f)
            for save_path in configs.values():
                if save_path and save_path.strip():
                    p = Path(save_path.strip()).resolve()
                    if p not in allowed_roots:
                        allowed_roots.append(p)
    except Exception:
        pass


def is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def clean_optional(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None

def parse_date(value: str | None) -> date | None:
    value = clean_optional(value)
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"日期格式不正确: {value}")


def main() -> None:
    import socket

    print("正在启动翻译报价系统...")

    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print(f"局域网访问地址: http://{local_ip}:5000/mobile")
    print(f"本机访问地址:   http://127.0.0.1:5000")

    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
