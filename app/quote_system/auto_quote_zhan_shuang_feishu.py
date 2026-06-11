"""战双飞书自动报价：双月汇总发行本地化沟通群需求报价单"""
from __future__ import annotations

import sys
import traceback
import shutil
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from quote_system.feishu_client import FeishuClient, excel_date_serial_to_date
from quote_system.memoq_html import parse_memoq_html
from quote_system.save_path_config import get_save_path
from quote_system.paths import get_quote_history_dir

SPREADSHEET_TOKEN = "Wup0wnUPIiiIr2k8T23c4zASnjd"
SHEET_ID = "WmpiFq"

LANG_CONFIG = {
    "中-韩": ("中韩", 0.64),
    "英-中": ("英译", 0.72),
    "日-韩": ("日韩", 0.72),
    "英-韩": ("英韩", 0.72),
}

TEMPLATE_TYPE_NAMES = [
    "All",
    "X-translated / double context",
    "Repetition",
    1.01,
    1,
    "95%-99%",
    "85%-94%",
    "75%-84%",
    "50%-74%",
    "No match",
]

# MemoQ HTML 实际输出可能使用这些别名
TYPE_NAME_ALIASES: dict[str | float, list[str]] = {
    "All": ["all", "全部", "すべて"],
    "X-translated / double context": [
        "x-translated / double context", "双重上下文", "クロス翻訳",
    ],
    "Repetition": ["repetition", "重复", "繰り返し"],
    1.01: ["101%", "101", "1.01"],
    1: ["100%", "100", "1"],
    "95%-99%": ["95%-99%", "95% - 99%", "95-99"],
    "85%-94%": ["85%-94%", "85% - 94%", "85-94"],
    "75%-84%": ["75%-84%", "75% - 84%", "75-84"],
    "50%-74%": ["50%-74%", "50% - 74%", "50-74"],
    "No match": ["no match", "一致しない", "无匹配"],
}


def get_bi_monthly_label(today: date) -> tuple[int, int, str]:
    m = today.month
    if m in (4, 5):
        return (4, 5, "4＆5")
    elif m in (6, 7):
        return (6, 7, "6＆7")
    elif m in (8, 9):
        return (8, 9, "8＆9")
    elif m in (10, 11):
        return (10, 11, "10＆11")
    elif m in (12, 1):
        return (12, 1, "12＆1")
    else:
        return (2, 3, "2＆3")


def get_bi_monthly_label_for_date(d: date) -> tuple[int, int, str]:
    """根据返回日期判断属于哪个双月"""
    m = d.month
    if m in (4, 5):
        return (4, 5, "4＆5")
    elif m in (6, 7):
        return (6, 7, "6＆7")
    elif m in (8, 9):
        return (8, 9, "8＆9")
    elif m in (10, 11):
        return (10, 11, "10＆11")
    elif m in (12, 1):
        return (12, 1, "12＆1")
    else:
        return (2, 3, "2＆3")


def feishu_quote_path(today: date) -> Path:
    _, _, label = get_bi_monthly_label(today)
    return get_quote_history_dir() / "库洛游戏" / "战双发行" / f"{today.year}年{label}月发行本地化沟通群需求报价单.xlsx"


def feishu_quote_path_for_date(d: date) -> Path:
    """根据返回日期生成对应的报价单路径"""
    _, _, label = get_bi_monthly_label_for_date(d)
    return get_quote_history_dir() / "库洛游戏" / "战双发行" / f"{d.year}年{label}月发行本地化沟通群需求报价单.xlsx"


def find_feishu_rows(rows: list[list]) -> list[dict]:
    result = []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        source = str(row[1]).strip() if len(row) > 1 else ""
        if source != "飞书":
            continue

        status = str(row[5]).strip() if len(row) > 5 else ""
        if status and status != "None":
            continue

        name = str(row[0]).strip() if len(row) > 0 else ""
        g_val = row[6] if len(row) > 6 else None
        if not g_val or not isinstance(g_val, list) or len(g_val) == 0:
            continue
        attach = g_val[0]
        file_token = attach.get("fileToken")
        file_name = attach.get("text", "")
        if not file_token:
            continue

        date_serial = row[2] if len(row) > 2 else None
        date_str = excel_date_serial_to_date(date_serial) if date_serial else ""
        deliv_serial = row[3] if len(row) > 3 else None
        deliv_str = excel_date_serial_to_date(deliv_serial) if deliv_serial else ""
        lang = str(row[4]).strip() if len(row) > 4 else ""

        result.append({
            "row_num": i + 1,
            "req_name": name,
            "lang": lang,
            "file_token": file_token,
            "file_name": file_name,
            "date_str": date_str,
            "deliv_str": deliv_str,
        })
    return result


def load_quoted_names(xlsx_path: Path) -> set[str]:
    """读取报价列表中已有的需求名（B列）"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    names = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v:
            names.add(str(v).strip())
    wb.close()
    return names


def calc_billable(stats, lang: str = "") -> int:
    total = 0
    zero_types = {"x-translated / double context", "repetition", "101%",
                  "100%", "95%-99%", "context", "exact"}
    is_en = lang == "英-韩"
    for row in stats.rows:
        t = str(row.type).strip().lower() if row.type else ""
        if t in zero_types or t in ("all",):
            continue
        if is_en:
            total += row.source_non_asian_words or 0
        else:
            total += row.source_asian_characters or 0
    return total


def _match_stats_row(stats_rows, template_name: str) -> object | None:
    """在解析出的统计行中查找与模板类型名匹配的行"""
    aliases = TYPE_NAME_ALIASES.get(template_name, [template_name])
    alias_set = {a.lower() for a in aliases}
    for sr in stats_rows:
        if str(sr.type).strip().lower() in alias_set:
            return sr
    return None


def append_one(client: FeishuClient, xlsx_path: Path, item: dict) -> int:
    """追加一条飞书需求到报价单，返回计费字数"""
    print(f"  处理: {item['req_name']}")

    # 下载HTML
    html_dir = ROOT / "outputs" / "zhan_shuang_html"
    html_dir.mkdir(parents=True, exist_ok=True)
    html_path = html_dir / item["file_name"]
    client.download_attachment(item["file_token"], html_path)

    try:
        stats = parse_memoq_html(html_path)
    except Exception:
        stats = None

    billable = calc_billable(stats, item["lang"]) if stats else 0
    print(f"  报价字数: {billable}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws_list = wb[wb.sheetnames[0]]  # 报价列表
    ws_stats = wb[wb.sheetnames[1]]  # Sheet1

    # 找最后一个有数据的行之后的位置
    last_data = 1
    for r in range(2, ws_list.max_row + 1):
        if ws_list.cell(r, 2).value:
            last_data = r
    next_row = last_data + 1

    seq = next_row - 1
    file_index = seq - 1  # 0-based
    block_start = 1 + file_index * 15

    # --- Sheet1 统计页 ---
    ws_stats.cell(block_start, 1).value = f"=需求列表!B{next_row}"

    # 统计表头（只有第一个block需要写，模板已有header row但新block需要写headers）
    # 实际上模板从 block 2 开始已经预填了所有 header row，不需要额外写

    # 写入统计行数据
    is_en_kr = item["lang"] == "英-韩"
    for idx, tmpl_type in enumerate(TEMPLATE_TYPE_NAMES):
        data_row = block_start + 2 + idx
        matched = _match_stats_row(stats.rows, tmpl_type) if stats else None

        ws_stats.cell(data_row, 1).value = tmpl_type
        if matched:
            ws_stats.cell(data_row, 2).value = matched.segments
            ws_stats.cell(data_row, 3).value = matched.source_words
            ws_stats.cell(data_row, 4).value = matched.source_non_asian_words
            # 英-韩：Source Asian characters为0，用non-Asian words填充
            if is_en_kr:
                ws_stats.cell(data_row, 5).value = matched.source_non_asian_words
            else:
                ws_stats.cell(data_row, 5).value = matched.source_asian_characters
            ws_stats.cell(data_row, 6).value = matched.source_chars
            ws_stats.cell(data_row, 7).value = matched.source_tags
            ws_stats.cell(data_row, 8).value = matched.percent

    # --- 报价列表 ---
    ws_list.cell(next_row, 1).value = seq
    ws_list.cell(next_row, 2).value = item["req_name"]

    if item["date_str"]:
        dt = datetime.strptime(item["date_str"], "%Y-%m-%d")
        c = ws_list.cell(next_row, 3)
        c.value = dt
        c.number_format = 'YYYY/MM/DD'

    if item["deliv_str"]:
        dt = datetime.strptime(item["deliv_str"], "%Y-%m-%d")
        c = ws_list.cell(next_row, 4)
        c.value = dt
        c.number_format = 'YYYY/MM/DD'

    lang_display, price = LANG_CONFIG.get(item["lang"], (item["lang"], 0.64))
    ws_list.cell(next_row, 5).value = lang_display
    ws_list.cell(next_row, 6).value = f"=Sheet1!J{block_start + 12}"
    ws_list.cell(next_row, 7).value = price
    c = ws_list.cell(next_row, 8)
    c.value = f"=F{next_row}*G{next_row}*1.06"
    c.number_format = '0.00'

    wb.save(xlsx_path)
    wb.close()

    if html_path.exists():
        html_path.unlink()

    return billable


def run() -> None:
    today = date.today()
    month1, month2, label = get_bi_monthly_label(today)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] 战双飞书{month1}&{month2}月报价自动化开始")

    client = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN, sheet_id=SHEET_ID)

    try:
        rows = client.read_sheet()
    except Exception as e:
        print(f"读取飞书表格失败: {e}")
        return

    all_feishu = find_feishu_rows(rows)
    if not all_feishu:
        print("没有找到飞书来源的需求")
        return

    # 按返回日期分组
    grouped_items: dict[str, list] = {}
    for item in all_feishu:
        if not item['deliv_str']:
            print(f"  跳过无返回日期的需求: {item['req_name']}")
            continue
        try:
            deliv_date = datetime.strptime(item['deliv_str'], "%Y-%m-%d").date()
        except ValueError:
            print(f"  跳过返回日期格式错误的需求: {item['req_name']} ({item['deliv_str']})")
            continue
        _, _, bi_label = get_bi_monthly_label_for_date(deliv_date)
        if bi_label not in grouped_items:
            grouped_items[bi_label] = []
        grouped_items[bi_label].append(item)

    if not grouped_items:
        print("没有找到有返回日期的需求")
        return

    print(f"按返回日期分组: {', '.join(f'{label}月({len(items)}条)' for label, items in grouped_items.items())}")

    # 只处理当前双月及之后的组，跳过已结算的历史月份
    cur_m1, cur_m2, cur_label = get_bi_monthly_label(today)

    def _bi_month_order(label: str) -> int:
        m = int(label.split("＆")[0])
        return m if m >= 2 else m + 12

    cur_order = _bi_month_order(cur_label)

    total_processed = 0
    for bi_label, items in grouped_items.items():
        if _bi_month_order(bi_label) < cur_order:
            print(f"跳过已结算的{bi_label}月({len(items)}条)")
            continue
        # 获取该组的报价单路径
        sample_date = datetime.strptime(items[0]['deliv_str'], "%Y-%m-%d").date()
        qpath = feishu_quote_path_for_date(sample_date)

        # 确保报价单文件存在
        if not qpath.exists():
            template = ROOT / "模板" / "报价单模板" / "2026年4＆5月发行本地化沟通群需求报价单模板.xlsx"
            if not template.exists():
                print(f"模板文件不存在: {template}")
                continue
            qpath.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(template, qpath)
            print(f"已从模板创建: {qpath.name}")

        # 找出尚未添加到报价单的需求
        existing = load_quoted_names(qpath)
        new_items = [it for it in items if it["req_name"] not in existing]

        if not new_items:
            print(f"{bi_label}月报价单: 所有需求已处理完毕")
            continue

        print(f"\n{bi_label}月报价单: 找到 {len(new_items)} 个新需求:")
        for it in new_items:
            print(f"  行{it['row_num']}: {it['req_name']} | {it['lang']} | {it['date_str']}")

        for item in new_items:
            row_num = item["row_num"]
            try:
                billable = append_one(client, qpath, item)
                # 更新状态和I列
                client.write_cell(row_num, "F", "已报价")
                client.write_cell(row_num, "I", billable)
                print(f"  行{row_num} 状态→已生成, I列→{billable}")

                try:
                    from settlement.settlement_tracker import quick_record, add_record
                    lang_display, price = LANG_CONFIG.get(item["lang"], (item["lang"], 0.64))
                    add_record(quick_record(
                        project_key="zhan_shuang_faxing",
                        company="库洛游戏",
                        req_name=item["req_name"],
                        word_count=billable,
                        total_price=round(billable * price * 1.06, 2),
                        quote_file=qpath.name,
                        language=lang_display,
                        delivery_date=item["deliv_str"] or None,
                        source="auto",
                        billable_words=billable,
                    ))
                except Exception:
                    pass
                total_processed += 1

            except Exception as e:
                print(f"处理行 {row_num} 失败: {e}")
                traceback.print_exc()

        print(f"\n{bi_label}月报价单已更新: {qpath.name}")

        save_path = get_save_path("zhan_shuang_faxing")
        if save_path:
            dest = Path(save_path) / qpath.name
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(qpath), str(dest))
            print(f"已另存至: {dest}")

    print(f"\n总计处理 {total_processed} 条需求")


if __name__ == "__main__":
    run()
