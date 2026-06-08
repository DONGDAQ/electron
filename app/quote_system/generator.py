from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .memoq_html import MemoqStats, parse_memoq_html, quote_words, normalize_label
from .projects import ProjectConfig, find_workspace_dirs


STAT_SHEET_BRACKET = "【字数统计】"
STAT_SHEET_ZHAN = "字数统计"
QUOTE_SHEET = "本地化报价"


@dataclass(frozen=True)
class QuoteRequest:
    project: ProjectConfig
    html_paths: list[Path]
    languages: list[str] | None
    quote_date: date
    delivery_date: date | None
    service_content: str | None
    request_name: str | None
    include_extract: bool
    output_path: Path | None = None
    price_overrides: dict[str, float] | None = None
    per_file_languages: list[list[str]] | None = None
    task_no: str | None = None
    delivery_time: str | None = None
    date_label: str | None = None


@dataclass(frozen=True)
class QuoteResult:
    output_path: Path
    stats: list[MemoqStats]
    final_path: Path


def generate_quote(root: Path, request: QuoteRequest) -> QuoteResult:
    template_dir, history_root = find_workspace_dirs(root)
    template_path = template_dir / request.project.template_file
    if not template_path.exists():
        raise FileNotFoundError(f"找不到模板: {template_path}")

    stats = [parse_memoq_html(path) for path in request.html_paths]
    final_path = default_output_path(history_root, request, stats)
    
    default_path = avoid_overwrite(final_path)
    default_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, default_path)

    wb = openpyxl.load_workbook(default_path)
    if request.project.generator == "zhan_shuang":
        fill_zhan_shuang(wb, request, stats)
    elif request.project.generator == "bang2":
        fill_bang2(wb, request, stats)
    elif request.project.generator == "huanta":
        fill_huan_ta(wb, request, stats)
    else:
        fill_punctuation_project(wb, request, stats)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(default_path)
    
    if request.output_path:
        custom_path = request.output_path / final_path.name
        custom_path = avoid_overwrite(custom_path)
        custom_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(default_path, custom_path)
        return QuoteResult(output_path=default_path, stats=stats, final_path=custom_path)
    
    return QuoteResult(output_path=default_path, stats=stats, final_path=default_path)


def fill_punctuation_project(wb, request: QuoteRequest, stats_list: list[MemoqStats]) -> None:
    quote_ws = wb[QUOTE_SHEET]
    default_languages = request.languages or list(request.project.default_languages)
    quote_ws["I7"] = datetime.combine(request.quote_date, datetime.min.time())
    quote_ws["C19"] = delivery_value(request.delivery_date)
    clear_quote_rows(quote_ws, 10, 15, columns=(3, 6, 7, 8, 10))

    quote_row = 10
    for index, stats in enumerate(stats_list, start=1):
        stat_ws = prepare_stats_sheet(wb, index, STAT_SHEET_BRACKET)
        write_a_to_h_stats(stat_ws, stats, start_row=1)
        service_content = request.service_content or default_service_content(stats)
        
        file_languages = request.per_file_languages[index - 1] if request.per_file_languages else default_languages

        for language in file_languages:
            price = get_price(request, language)
            words = f"='{stat_ws.title}'!F3" if language == "摘字" else f"='{stat_ws.title}'!S6"
            quote_row = write_quote_line(
                quote_ws,
                quote_row,
                service_content,
                words,
                language,
                price,
                use_k_formula=True,
            )


def fill_bang2(wb, request: QuoteRequest, stats_list: list[MemoqStats]) -> None:
    quote_ws = wb[QUOTE_SHEET]
    default_languages = request.languages or list(request.project.default_languages)
    quote_ws["I7"] = request.quote_date.strftime("%Y/%m/%d")
    quote_ws["C25"] = delivery_value(request.delivery_date)
    clear_quote_rows(quote_ws, 10, 21, columns=(3, 6, 7, 8, 9, 10, 11))

    quote_row = 10
    for index, stats in enumerate(stats_list, start=1):
        stat_ws = prepare_stats_sheet(wb, index, STAT_SHEET_BRACKET)
        write_a_to_h_stats(stat_ws, stats, start_row=1)
        service_content = request.service_content or default_service_content(stats)
        words_formula = quote_words(stats)
        
        file_languages = request.per_file_languages[index - 1] if request.per_file_languages else default_languages

        for language in file_languages:
            price = get_price(request, language)
            words = stats.all_row.source_chars or 0 if language == "摘字" else words_formula
            quote_row = write_quote_line(
                quote_ws,
                quote_row,
                service_content,
                words,
                language,
                price,
                use_k_formula=False,
            )
    
    for row in range(10, 22):
        cell_value = quote_ws.cell(row, 3).value
        quote_ws.row_dimensions[row].hidden = cell_value is None or str(cell_value).strip() == ""


def fill_zhan_shuang(wb, request: QuoteRequest, stats_list: list[MemoqStats]) -> None:
    if len(stats_list) > 6:
        raise ValueError("战双模板最多支持 6 个需求")
    quote_ws = wb[QUOTE_SHEET]
    stat_ws = wb[STAT_SHEET_ZHAN]
    
    quote_ws["H7"] = request.quote_date.strftime("%Y/%m/%d")
    
    for index, stats in enumerate(stats_list):
        quote_row = 10 + index
        start_row = 6 + index * 16
        
        if index == 0:
            quote_ws.cell(quote_row, 3).value = request.service_content or default_service_content(stats)
        else:
            quote_ws.cell(quote_row, 3).value = default_service_content(stats)
        
        write_b_to_h_stats(stat_ws, stats, start_row=start_row)
        quote_ws.cell(quote_row, 6).value = f"={STAT_SHEET_ZHAN}!J{start_row + 10}"
        quote_ws.cell(quote_row, 7).value = get_price(request, "中译韩")
        quote_ws.cell(quote_row, 8).value = f"=F{quote_row}*G{quote_row}"
        quote_ws.cell(quote_row, 9).value = 0.06
        quote_ws.cell(quote_row, 10).value = f"=H{quote_row}*106%"
    
    for row in range(10, 16):
        cell_value = quote_ws.cell(row, 3).value
        quote_ws.row_dimensions[row].hidden = cell_value is None or str(cell_value).strip() == ""
    
    if request.delivery_date:
        quote_ws["F20"] = request.delivery_date.strftime("%Y/%m/%d")


def fill_huan_ta(wb, request: QuoteRequest, stats_list: list[MemoqStats]) -> None:
    """幻塔模板：一个报价单对应一个HTML文件。模板有3个sheet：本地化报价、字数统计、具体数据"""
    if len(stats_list) != 1:
        raise ValueError("幻塔模板一次只支持一个HTML文件")

    stats = stats_list[0]
    data_ws = wb["具体数据"]
    quote_ws = wb["本地化报价"]

    data_ws["A1"] = stats.title or stats.source_path.name

    match_map = _build_huanta_match_map(stats.rows)
    all_row = stats.all_row

    _write_row(data_ws, 9, all_row.as_a_to_h())

    # Rows 10-18: mapping of match types to template rows
    template_rows = {
        "context": 10,
        "exact": 11,
        "101": 12,
        "100": 13,
        "95-99": 14,
        "85-94": 15,
        "75-84": 16,
        "50-74": 17,
        "no_match": 18,
    }

    written_rows = set()
    for row_type, template_row in template_rows.items():
        matched = match_map.get(row_type)
        if matched:
            _write_row(data_ws, template_row, matched.as_a_to_h())
            written_rows.add(template_row)

    # 清除模板中没有匹配数据的行
    for template_row in range(10, 19):
        if template_row not in written_rows:
            for col in range(1, 9):
                data_ws.cell(template_row, col).value = None

    # 填写报价单信息
    req_name = request.request_name or default_service_content(stats)
    content_type = request.languages[0] if request.languages else "翻译+润色"

    if request.task_no:
        quote_ws["C10"] = f"{req_name}\n{request.task_no}"
    else:
        quote_ws["C10"] = req_name

    price = get_price(request, content_type)
    quote_ws["G10"] = price

    # 报价单号 E7: 委托日期 + 同天序号后缀
    if request.date_label:
        quote_ws["E7"] = request.date_label
    else:
        quote_ws["E7"] = request.quote_date.strftime("%Y%m%d")
    # 报价日期 H7
    from datetime import datetime
    quote_ws["H7"] = datetime.combine(request.quote_date, datetime.min.time())
    # 交付时间 C15
    if request.delivery_date:
        delivery_str = request.delivery_date.strftime("%Y-%m-%d")
        if request.delivery_time:
            delivery_str += f" {request.delivery_time}"
        quote_ws["C15"] = f"交付时间  {delivery_str}"


MATCH_PATTERNS = {
    "context": {"上下文", "context", "コンテキスト", "クロスファイル", "クロス翻訳", "二重コンテキスト",
                "cross file", "跨文件", "x 翻译/双重上下文"},
    "exact": {"完全一致", "完全匹配", "exact", "完全一致", "重复", "繰り返し", "repetition"},
    "101": {"101%", "101", "1.01"},
    "100": {"100%", "100", "1"},
    "95-99": {"95%-99%", "95% - 99%", "95-99"},
    "85-94": {"85%-94%", "85% - 94%", "85-94"},
    "75-84": {"75%-84%", "75% - 84%", "75-84"},
    "50-74": {"50%-74%", "50% - 74%", "50-74"},
    "no_match": {"一致しない", "无匹配", "no match", "0%-49%", "0% - 49%", "0-49"},
}


def _build_huanta_match_map(rows: list) -> dict:
    result = {}
    for row in rows:
        normalized = normalize_label(row.type)
        for key, aliases in MATCH_PATTERNS.items():
            if key in result:
                continue
            if normalized in {normalize_label(a) for a in aliases}:
                result[key] = row
                break
    return result


def _write_row(ws, row_num: int, values: list):
    for col, value in enumerate(values, start=1):
        ws.cell(row_num, col).value = value


def prepare_stats_sheet(wb, index: int, base_name: str) -> Worksheet:
    if index == 1:
        return wb[base_name]
    name = f"{base_name}{index}"
    if name in wb.sheetnames:
        return wb[name]
    copied = wb.copy_worksheet(wb[base_name])
    copied.title = name
    return copied


def write_a_to_h_stats(ws: Worksheet, stats: MemoqStats, start_row: int) -> None:
    ws.cell(start_row, 1).value = stats.title or stats.source_path.name
    headers = [
        "Type",
        "Segments",
        "Source words",
        "Source non-Asian words",
        "Source Asian characters",
        "Source chars",
        "Source tags",
        "Percent",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(start_row + 1, col).value = value
    for offset, row in enumerate(stats.rows, start=2):
        excel_row = start_row + offset
        values = row.as_a_to_h()
        for col, value in enumerate(values, start=1):
            ws.cell(excel_row, col).value = value


def write_b_to_h_stats(ws: Worksheet, stats: MemoqStats, start_row: int) -> None:
    for offset, row in enumerate(stats.rows[:11], start=0):
        excel_row = start_row + offset
        values = row.as_b_to_h()
        for col, value in enumerate(values, start=2):
            ws.cell(excel_row, col).value = value


def clear_quote_rows(ws: Worksheet, start_row: int, end_row: int, columns: tuple[int, ...]) -> None:
    for row in range(start_row, end_row + 1):
        for col in columns:
            ws.cell(row, col).value = None


def write_quote_line(
    ws: Worksheet,
    row: int,
    service_content: str,
    words: int | str,
    language: str,
    price: float,
    *,
    use_k_formula: bool,
) -> int:
    ws.cell(row, 3).value = service_content
    ws.cell(row, 6).value = words
    ws.cell(row, 7).value = language
    ws.cell(row, 8).value = price
    if use_k_formula:
        ws.cell(row, 9).value = f"=K{row}/1.06"
        ws.cell(row, 10).value = 0.06
        ws.cell(row, 11).value = f"=H{row}*F{row}"
    else:
        ws.cell(row, 9).value = f"=F{row}*H{row}"
        ws.cell(row, 10).value = 0.06
        ws.cell(row, 11).value = f"=I{row}*1.06"
    return row + 1


def ensure_price(project: ProjectConfig, language: str) -> None:
    if language not in project.prices:
        supported = "、".join(project.prices)
        raise ValueError(f"{project.display_name} 不支持语种/服务: {language}。支持: {supported}")


def get_price(request: QuoteRequest, language: str) -> float:
    if request.price_overrides and language in request.price_overrides:
        return request.price_overrides[language]
    ensure_price(request.project, language)
    return request.project.prices[language]


def delivery_value(delivery_date: date | None) -> datetime | str:
    if delivery_date is None:
        return "未定"
    return datetime.combine(delivery_date, datetime.min.time())


def default_output_path(history_root: Path, request: QuoteRequest, stats: list[MemoqStats]) -> Path:
    project_dir = history_root / request.project.history_dir
    
    safe_name = sanitize_filename(request.request_name or stats[0].source_path.stem)
    if request.project.key == "zhan_shuang":
        source_text = request.request_name or stats[0].source_path.stem
        request_part = extract_zhan_request_name(source_text)
        filename = f"{request.project.file_prefix} {request_part}.xlsx"
    elif request.project.key == "umamusume":
        filename = f"报价单_《代号PD》{request.quote_date:%m%d}需求.xlsx"
    elif request.project.key == "bang2":
        filename = f"报价单_《BANG2》{request.quote_date:%m%d}需求.xlsx"
    elif request.project.key == "hbr":
        filename = f"报价单_《炽焰天穹》{request.quote_date:%m%d}需求.xlsx"
    elif request.project.generator == "huanta":
        safe_name = sanitize_filename(request.request_name or stats[0].source_path.stem)
        if safe_name.lower().endswith(".xlsx"):
            safe_name = safe_name[:-5]
        filename = f"报价_{safe_name}.xlsx"
    else:
        filename = f"{request.project.file_prefix}_{safe_name}.xlsx"
    
    filename = filename.replace("报价单", "报价单^")
    
    return project_dir / filename


def avoid_overwrite(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"无法生成不重名文件: {path}")


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "quote"


def extract_zhan_request_name(value: str) -> str:
    match = re.search(r"(【?战双v?[\d.]+】?[^/\t\n]+?)\s*(?:[\t/ ]|$)", value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r"(【?战双v?[\d.]+】?.*)", value, flags=re.IGNORECASE)
    return match.group(1).strip() if match else value


def display_stem(path: Path) -> str:
    stem = path.stem
    for suffix in ("_字数统计", " 字数统计", "-字数统计"):
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def default_service_content(stats: MemoqStats) -> str:
    candidates = [stats.title, stats.source_path.stem]
    for candidate in candidates:
        cleaned = clean_memoq_stat_title(candidate)
        if cleaned:
            return cleaned
    return display_stem(stats.source_path)


def clean_memoq_stat_title(value: str | None) -> str:
    if not value:
        return ""
    text = value.strip()
    file_match = re.search(r"ファイル\s*(?:\[[^\]]+\])?\s*(.+?)\s*の統計$", text)
    if file_match:
        text = file_match.group(1).strip()
    text = re.sub(r"^Statistics\s+for\s+file\(s\)\s*(?:\[[^\]]+\])?\s*", "", text, flags=re.IGNORECASE)
    for suffix in (".xlsx", ".xlsm", ".xls", ".html", ".htm"):
        if text.lower().endswith(suffix):
            text = text[: -len(suffix)]
            break
    for suffix in ("_字数统计", " 字数统计", "-字数统计"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    return text.strip()



