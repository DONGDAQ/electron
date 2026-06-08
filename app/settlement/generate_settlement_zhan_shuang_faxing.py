"""战双发行月度结算单生成"""
import sys
import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from quote_system.feishu_client import FeishuClient
from quote_system.paths import get_settlement_dir

SPREADSHEET_TOKEN = "Wup0wnUPIiiIr2k8T23c4zASnjd"
SHEET_ID = "WmpiFq"

UNIT_PRICES = {
    "中译韩": 0.64,
    "中-韩": 0.64,
    "日译韩": 0.72,
    "日-韩": 0.72,
    "英译韩": 0.72,
    "英-韩": 0.72,
}

BASE_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATE_DIR = BASE_DIR / "app" / "模板" / "结算模板"


def excel_serial_to_date(serial):
    base = datetime(1899, 12, 30)
    return base + timedelta(days=float(serial))


def read_feishu_data(year, months):
    client = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN, sheet_id=SHEET_ID)
    data = client.read_sheet()

    month_serials = []
    base = datetime(1899, 12, 30)
    for m in months:
        month_start = datetime(year, m, 1)
        if m == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, m + 1, 1)
        month_serials.append((
            (month_start - base).days,
            (month_end - base).days,
        ))

    records = []
    for i, row in enumerate(data):
        if i == 0:
            continue
        if len(row) < 9:
            continue

        source = str(row[1] or "").strip()
        if source != "飞书":
            continue

        status = str(row[5] or "").strip()
        if status != "已报价":
            continue

        deliv_val = row[3] if len(row) > 3 else None
        if deliv_val is None:
            continue
        try:
            deliv_serial = float(deliv_val)
        except (TypeError, ValueError):
            continue

        in_range = False
        for s_start, s_end in month_serials:
            if s_start <= deliv_serial < s_end:
                in_range = True
                break
        if not in_range:
            continue

        language = str(row[4] or "").strip() if len(row) > 4 else ""
        word_count = row[8] if len(row) > 8 else None
        entrust_date = row[2] if len(row) > 2 else None

        records.append({
            'row_index': i + 1,
            'file_name': str(row[0] or ""),
            'language': language,
            'entrust_date': float(entrust_date) if entrust_date is not None else None,
            'deliv_date': deliv_serial,
            'word_count': float(word_count) if word_count is not None else 0,
            'unit_price': UNIT_PRICES.get(language, 0.64),
        })

    return records


def _month_label(months):
    return "&".join(str(m) for m in months)


def generate_settlement_excel(records, year, months, output_dir):
    month_label = _month_label(months)
    template = TEMPLATE_DIR / "【大连游者之家】结算单_发行部韩语本地化需求26年2&3月模板.xlsx"
    filename = f"【大连游者之家】结算单_发行部韩语本地化需求{year}年{month_label}月.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime('%H%M%S')
            filename = f"【大连游者之家】结算单_发行部韩语本地化需求{year}年{month_label}月_{ts}.xlsx"
            output_path = output_dir / filename

    shutil.copy2(str(template), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    ws['J7'] = datetime.now()

    data_start_row = 10

    for i, rec in enumerate(records):
        r = data_start_row + i
        ws.cell(row=r, column=3, value=rec['file_name'])
        ws.cell(row=r, column=5, value=rec['language'])
        if rec['entrust_date'] is not None:
            ws.cell(row=r, column=6, value=rec['entrust_date'])
            ws.cell(row=r, column=6).number_format = 'm/d'
        ws.cell(row=r, column=7, value=rec['deliv_date'])
        ws.cell(row=r, column=7).number_format = 'm/d'
        if rec['word_count']:
            ws.cell(row=r, column=8, value=rec['word_count'])
        ws.cell(row=r, column=9, value=rec['unit_price'])

    last_data_row = data_start_row + len(records) - 1
    total_row = None
    for row in range(last_data_row + 1, ws.max_row + 1):
        for col in range(1, 19):
            val = ws.cell(row=row, column=col).value
            if val and '合' in str(val):
                total_row = row
                break
        if total_row:
            break
    if total_row:
        for r in range(last_data_row + 1, total_row):
            ws.row_dimensions[r].hidden = True

    wb.save(str(output_path))
    print(f'[OK] 结算单已保存: {output_path}')
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description='战双发行月度结算单生成')
    parser.add_argument('--year', type=int, default=2026)
    parser.add_argument('--months', type=str, default='5')
    parser.add_argument('--output', type=str, default=None)
    args = parser.parse_args()

    year = args.year
    months = [int(m.strip()) for m in args.months.split(',')]
    month_label = _month_label(months)
    output_dir = Path(args.output) if args.output else get_settlement_dir() / f"{year}年{months[-1]}月" / "库洛游戏" / "战双发行"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f'读取飞书数据: {year}年{month_label}月...')
    records = read_feishu_data(year, months)
    print(f'找到 {len(records)} 条交付记录')

    if not records:
        print('没有找到匹配的交付记录，退出。')
        return

    generate_settlement_excel(records, year, months, output_dir)
    print('完成！')


if __name__ == '__main__':
    main()
