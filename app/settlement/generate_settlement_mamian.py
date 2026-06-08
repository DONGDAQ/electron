"""Bilibili 项目月度账单生成：从报价单历史 xlsx 读取数据 → 填充模板 → 移入已结算。
支持马娘(umamusume)、邦邦2(bang2)、炽焰天穹(hbr)。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from quote_system.paths import get_quote_history_dir, get_settlement_dir

TEMPLATE_DIR = ROOT / "模板" / "结算模板"
OUTPUT_DIR = get_settlement_dir()
SETTLEMENT_COMPANY = "Bilibili"
TEMPLATE_NAME = "大连游者之家翻译账单-哔哩哔哩游戏【代号PD】-2026年4月模板.xlsx"

BILL_CONFIG = {
    "umamusume": {
        "project": "马娘",
        "project_full_name": "优俊少女",
        "code": "代号PD",
        "quote_history_dir": "马娘",
    },
    "bang2": {
        "project": "bang2",
        "project_full_name": "邦邦2",
        "code": "邦邦2",
        "quote_history_dir": "bang2",
    },
    "hbr": {
        "project": "HBR",
        "project_full_name": "炽焰天穹",
        "code": "HBR",
        "quote_history_dir": "HBR",
    },
}


def _get_bill_config(project_key: str) -> dict:
    cfg = BILL_CONFIG.get(project_key)
    if not cfg:
        raise ValueError(f"不支持的账单项目: {project_key}")
    return cfg


def excel_serial_to_date(serial: float) -> date:
    return date(1899, 12, 30) + __import__("datetime").timedelta(days=serial)


def _sum_data_column(qs, col: int, start_row: int, end_row: int) -> int:
    total = 0
    for r in range(start_row, end_row):
        v = qs.cell(row=r, column=col).value
        if isinstance(v, (int, float)):
            total += int(round(v))
    return total


def _recalc_quote(xlsx_path: Path) -> bool:
    """用 Excel 打开报价单重算公式，保存缓存值。成功返回 True。"""
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        wb.Save()
        wb.Close()
        return True
    except Exception:
        return False
    finally:
        try:
            if wb:
                wb.Close()
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _read_summary_row(qs, project_key: str) -> float | None:
    """从报价 sheet 的合计行 K 列读取总金额。"""
    from decimal import Decimal, ROUND_HALF_UP
    for r in range(15, qs.max_row + 1):
        c3 = qs.cell(row=r, column=3).value
        if c3 and ("合计" in str(c3) or "合" == str(c3).strip()):
            val = qs.cell(row=r, column=11).value
            if isinstance(val, (int, float)) and val > 0:
                d = Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                return float(d)
            return None
    return None


def extract_quote_data(xlsx_path: Path, project_key: str = "umamusume") -> dict | None:
    """从报价单 xlsx 直接读取合计行 K 列总金额和 F 列总字数。无缓存则调 Excel 重算。"""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return None

    qs_name = None
    for sn in wb.sheetnames:
        s = sn.strip().replace(" ", "")
        if "报价" in s or "本地化" in s:
            qs_name = sn
            break
    if not qs_name:
        return None

    qs = wb[qs_name]

    quote_date = _parse_date(qs.cell(row=7, column=9).value)
    delivery_date = _find_delivery_date(qs)
    req_name = str(qs.cell(row=10, column=3).value or "").strip()

    total_from_summary = _read_summary_row(qs, project_key)
    if total_from_summary is None:
        _recalc_quote(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        qs = wb[qs_name]
        total_from_summary = _read_summary_row(qs, project_key)

    if total_from_summary is None:
        return None

    total_words = _sum_data_column(qs, 6, 10, 16)
    if total_words == 0:
        return None

    return {
        "req_name": req_name,
        "word_count": total_words,
        "total_price": total_from_summary,
        "quote_date": quote_date or "",
        "delivery_date": delivery_date or "",
    }


def _find_delivery_date(ws) -> str | None:
    """在报价 sheet 中查找交付日期：找含"交付"的标签，取下方单元格；没有则从 C19 读取。"""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=3).value
        if isinstance(val, str) and "交付" in val:
            date_val = ws.cell(row=row + 1, column=3).value
            parsed = _parse_date(date_val)
            if parsed:
                return parsed
    return _parse_date(ws.cell(row=19, column=3).value)


def _parse_date(val) -> str | None:
    if isinstance(val, (int, float)) and val > 40000:
        try:
            return excel_serial_to_date(val).isoformat()
        except Exception:
            pass
    elif isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    elif isinstance(val, date):
        return val.isoformat()
    elif isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def scan_quotes(target_year: int, target_month: int, project_key: str = "umamusume"):
    """扫描报价单历史文件夹（仅根目录），筛选指定月份的未结算记录。"""
    cfg = _get_bill_config(project_key)
    quote_dir = get_quote_history_dir() / SETTLEMENT_COMPANY / cfg["quote_history_dir"]
    if not quote_dir.exists():
        print(f"目录不存在: {quote_dir}")
        return []

    xlsx_files = sorted(quote_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)

    results = []
    for fpath in xlsx_files:
        data = extract_quote_data(fpath, project_key)
        if data is None:
            print(f"  [跳过] 无法读取: {fpath.name}")
            continue

        dd = data.get("delivery_date", "")
        if not dd:
            print(f"  [跳过] 无交付日期: {fpath.name}")
            continue

        try:
            dt = date.fromisoformat(dd)
        except (ValueError, TypeError):
            print(f"  [跳过] 日期格式错误: {fpath.name} -> {dd}")
            continue

        if dt.year == target_year and dt.month == target_month:
            data["filename"] = fpath.name
            results.append(data)

    return results


def move_settled(records: list[dict], year: int, month: int, project_key: str = "umamusume") -> Path:
    """将已结算的报价单文件移入 已结算/年月/ 目录。"""
    cfg = _get_bill_config(project_key)
    qdir = get_quote_history_dir() / SETTLEMENT_COMPANY / cfg["quote_history_dir"]
    settled_dir = qdir / "已结算" / f"{year}年{month}月"
    settled_dir.mkdir(parents=True, exist_ok=True)
    for rec in records:
        src = qdir / rec["filename"]
        if src.exists():
            shutil.move(str(src), str(settled_dir / rec["filename"]))
    return settled_dir


def main():
    parser = argparse.ArgumentParser(description="Bilibili 项目月度账单生成")
    parser.add_argument("--year", type=int, default=2026, help="结算年份")
    parser.add_argument("--month", type=int, required=True, help="结算月份")
    parser.add_argument("--project", type=str, default="umamusume", help="项目: umamusume/bang2/hbr")
    parser.add_argument("--dry-run", action="store_true", help="只预览不生成")
    args = parser.parse_args()

    year, month, pk = args.year, args.month, args.project
    cfg = _get_bill_config(pk)
    print(f"扫描 报价单历史/{SETTLEMENT_COMPANY}/{cfg['quote_history_dir']} ...")
    records = scan_quotes(year, month, pk)

    if not records:
        print(f"\n{year}年{month}月 没有找到未结算的 {cfg['project_full_name']} 交付记录。")
        return

    print(f"\n找到 {len(records)} 条未结算记录:")
    total_words = 0
    total_amount = 0.0
    for r in records:
        total_words += r["word_count"]
        total_amount += r["total_price"]
        dd = r.get("delivery_date", "?")
        print(f"  {r['filename']:45s}  {r['word_count']:>6}字  {r['total_price']:>10.2f}元  交付{dd}")

    print(f"  {'─' * 70}")
    print(f"  {'合计':>45s}  {total_words:>6}字  {total_amount:>10.2f}元")

    if args.dry_run:
        print("\n[dry-run] 不生成文件。")
        return

    gen = MamianSettlementGenerator(year, month, pk)
    output = gen.generate(records, args)
    print(f"\n账单已生成: {output}")

    settled_dir = move_settled(records, year, month, pk)
    print(f"已移入 {settled_dir}")


class MamianSettlementGenerator:
    def __init__(self, year: int, month: int, project_key: str = "umamusume"):
        self.year = year
        self.month = month
        self.project_key = project_key
        self.cfg = _get_bill_config(project_key)
        self.template_path = TEMPLATE_DIR / TEMPLATE_NAME

    def generate(self, records: list[dict], args) -> Path:
        template = self.template_path
        if not template.exists():
            raise FileNotFoundError(f"找不到模板: {template}")

        code = self.cfg["code"]
        project_name = self.cfg["project"]
        filename = f"大连游者之家翻译账单-哔哩哔哩游戏【{code}】-{self.year}年{self.month}月.xlsx"
        output_dir = OUTPUT_DIR / f"{self.year}年{self.month}月" / SETTLEMENT_COMPANY / project_name
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / filename

        if output_path.exists():
            stem = output_path.stem
            for idx in range(1, 100):
                alt = output_path.with_name(f"{stem}_{idx}.xlsx")
                if not alt.exists():
                    output_path = alt
                    break

        shutil.copy2(str(template), str(output_path))
        wb = openpyxl.load_workbook(str(output_path))
        ws = wb.active

        # 按交付日期排序
        records.sort(key=lambda r: r.get("delivery_date", ""))

        # 报价单名去掉"报价单"前缀及后面的分隔符
        for rec in records:
            rec["display_name"] = re.sub(r'^报价单[_\^]+', '', rec["filename"])

        data_start = 18
        template_slots = 9       # 模板预留数据行 18-26
        template_data_end = 27   # 模板小计行

        num_records = len(records)

        # 记录超出模板预留行数时插入空行
        if num_records > template_slots:
            extra = num_records - template_slots
            ws.insert_rows(template_data_end, extra)
            template_data_end += extra

        # 填写数据行
        for i, rec in enumerate(records):
            row = data_start + i
            ws.cell(row=row, column=2).value = i + 1
            ws.cell(row=row, column=3).value = rec["display_name"]
            ws.cell(row=row, column=4).value = code
            ws.cell(row=row, column=5).value = rec["word_count"]
            ws.cell(row=row, column=6).value = "邮件/群"
            ws.cell(row=row, column=7).value = rec.get("quote_date", "")
            ws.cell(row=row, column=8).value = rec.get("delivery_date", "")
            ws.cell(row=row, column=9).value = rec["total_price"]

        # 合计行
        total_row = data_start + num_records
        # 从模板小计行（原 27 行）复制样式到合计行
        style_src = template_data_end
        for col in range(2, 10):
            src_cell = ws.cell(row=style_src, column=col)
            dst_cell = ws.cell(row=total_row, column=col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = copy(src_cell.number_format)
                dst_cell.alignment = copy(src_cell.alignment)
        ws.cell(row=total_row, column=2).value = "合计"
        ws.cell(row=total_row, column=9).value = f"=SUM(I{data_start}:I{data_start + num_records - 1})"

        # 清空并隐藏多余的模板行（避免有框无内容）
        for r in range(total_row + 1, template_data_end + 1):
            for col in range(2, 10):
                ws.cell(row=r, column=col).value = None
            ws.row_dimensions[r].hidden = True

        # 更新 GRAND TOTAL 公式
        grand_total_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=8).value
            if v and "GRAND TOTAL" in str(v):
                grand_total_row = r
                break
        if grand_total_row:
            ws.cell(row=grand_total_row, column=9).value = f"=I{total_row}"
            for r in range(grand_total_row - 2, grand_total_row):
                cell = ws.cell(row=r, column=8).value
                if cell and "TOTAL PRICE" in str(cell):
                    ws.cell(row=r, column=9).value = f"=I{grand_total_row}/1.06"
                elif cell and "TAX" in str(cell):
                    ws.cell(row=r, column=9).value = f"=I{grand_total_row}-I{r-1}"

        # 更新日期为当前日期
        today = datetime.now()
        ws.cell(row=4, column=3).value = f"*本结算单生成日期：{today.year}年{today.month}月{today.day}日"

        wb.save(str(output_path))
        return output_path


if __name__ == "__main__":
    main()
