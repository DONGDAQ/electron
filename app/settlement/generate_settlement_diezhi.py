"""叠纸3项目（恋与深空/闪暖/ニキ新作）月度结算单生成"""
import os
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
import win32com.client

BASE_DIR = Path(__file__).resolve().parent.parent.parent
import sys as _sys
_sys.path.insert(0, str(BASE_DIR / "app"))
from quote_system.paths import get_quote_history_dir, get_settlement_dir

QUOTE_BASE = get_quote_history_dir()
TEMPLATE_PATH = BASE_DIR / "app" / "模板" / "结算模板" / "【大连游者之家】结算单_《X3.X6.闪暖》2026年4月模板.xlsx"
OUTPUT_BASE = get_settlement_dir()

PROJECTS = {
    'liandishenkong': {
        'display': '恋与深空',
        'short': 'X3',
        'dir': '叠纸/恋与深空',
        'delivery_cell': 'E39',
        'word_col': 6,
        'amount_col': 8,
    },
    'shining_nikki': {
        'display': '闪暖',
        'short': '闪暖',
        'dir': '叠纸/闪暖',
        'delivery_cell': 'E49',
        'word_col': 6,
        'amount_col': 8,
    },
    'niki_xinzuo': {
        'display': 'ニキ新作',
        'short': 'X6',
        'dir': '叠纸/ニキ新作',
        'delivery_cell': 'E51',
        'word_col': 7,
        'amount_col': 9,
    },
}

INVOICE_COMPANIES = {
    'shining_nikki': {
        'name': '芜湖叠纸网络科技有限公司',
        'tax_id': '91340207MA2MWRK069',
        'bank': '招商银行股份有限公司芜湖开发区支行',
        'account': '121933935610799',
        'address': '安徽省芜湖市鸠江区官陡街道鸠江北路77号芜湖广告产业园文化创意综合楼',
        'phone': '0553-8351236',
        'label': '闪暖',
    },
    'niki_xinzuo': {
        'name': '上海暖叠网络科技有限公司',
        'tax_id': '91310110MA1G99LX09',
        'bank': '招商银行上海联洋支行',
        'account': '121940336310803',
        'address': '上海市杨浦区政高路38号201室',
        'phone': '021-55390023',
        'label': 'X6',
    },
    'liandishenkong': {
        'name': '上海叠纸互娱网络科技有限公司',
        'tax_id': '91310110MA1G9BM29D',
        'bank': '招商银行上海分行联洋支行',
        'account': '121941875910401',
        'address': '上海市杨浦区政高路38号502室',
        'phone': '021-55390023',
        'label': 'X3',
    },
}


def _ensure_formulas_cached(filepath):
    import pythoncom
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(filepath)
        excel.CalculateUntilAsyncQueriesDone()
        wb.Save()
        wb.Close()
        excel.Quit()
    except Exception as e:
        print(f'[WARN] Excel重算失败: {e}')
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def scan_quote_files(history_dir):
    files = []
    if not history_dir.exists():
        return files
    for f in history_dir.iterdir():
        if f.is_file() and f.suffix == '.xlsx' and not f.name.startswith('~$'):
            files.append(f)
    settled = history_dir / "已结算"
    if settled.exists():
        for sub in settled.iterdir():
            if sub.is_dir():
                for f in sub.iterdir():
                    if f.is_file() and f.suffix == '.xlsx' and not f.name.startswith('~$'):
                        files.append(f)
    return sorted(files, key=lambda x: x.name)


def read_quote_data(quote_path, word_col, amount_col, delivery_cell):
    wb = openpyxl.load_workbook(quote_path, data_only=True)
    home_ws = wb[wb.sheetnames[0]]
    delivery_date = home_ws[delivery_cell].value

    total_words = 0.0
    total_amount = 0.0
    for row in home_ws.iter_rows(min_row=1, max_row=home_ws.max_row):
        c_val = row[2].value
        if c_val and '合' in str(c_val):
            total_words = float(row[word_col - 1].value or 0)
            total_amount = float(row[amount_col - 1].value or 0)
            break
    wb.close()

    if total_words == 0 and total_amount == 0:
        _ensure_formulas_cached(str(quote_path))
        wb = openpyxl.load_workbook(quote_path, data_only=True)
        home_ws = wb[wb.sheetnames[0]]
        delivery_date = home_ws[delivery_cell].value
        for row in home_ws.iter_rows(min_row=1, max_row=home_ws.max_row):
            c_val = row[2].value
            if c_val and '合' in str(c_val):
                total_words = float(row[word_col - 1].value or 0)
                total_amount = float(row[amount_col - 1].value or 0)
                break
        wb.close()

    return {
        'file_name': quote_path.name,
        'file_path': str(quote_path),
        'delivery_date': delivery_date,
        'total_words': total_words,
        'total_amount': total_amount,
    }


def read_x3_quote_by_type(quote_path, delivery_cell):
    """读取X3报价单，按I列备注分类统计字数"""
    def _read_x3_data(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        delivery_date = ws[delivery_cell].value
        type_words = {"TEP": 0.0, "润色/我方初翻": 0.0, "润色/非我方初翻": 0.0}
        for r in range(10, 36):
            f_val = float(ws.cell(r, 6).value or 0)
            if f_val == 0:
                continue
            note = str(ws.cell(r, 9).value or "").strip()
            if note == "TEP":
                type_words["TEP"] += f_val
            elif note == "润色/我方初翻":
                type_words["润色/我方初翻"] += f_val
            elif note == "润色/非我方初翻":
                type_words["润色/非我方初翻"] += f_val
            elif note == "TEP润色/我方初翻":
                type_words["TEP"] += f_val
                type_words["润色/我方初翻"] += f_val
            elif note == "TEP润色/非我方初翻":
                type_words["TEP"] += f_val
                type_words["润色/非我方初翻"] += f_val
        wb.close()
        return delivery_date, type_words

    delivery_date, type_words = _read_x3_data(quote_path)
    if not any(type_words.values()):
        _ensure_formulas_cached(str(quote_path))
        delivery_date, type_words = _read_x3_data(quote_path)

    return {
        'file_name': quote_path.name,
        'delivery_date': delivery_date,
        'type_words': {k: v for k, v in type_words.items() if v > 0},
    }


X3_PRICES = {"TEP": 0.52, "润色/我方初翻": 0.26, "润色/非我方初翻": 0.4}


def find_quotes_for_month(history_dir, word_col, amount_col, delivery_cell, year, month):
    results = []
    for f in scan_quote_files(history_dir):
        try:
            data = read_quote_data(f, word_col, amount_col, delivery_cell)
        except Exception as e:
            print(f'[WARN] 读取报价单失败: {f.name}: {e}')
            continue
        dd = data['delivery_date']
        if not isinstance(dd, datetime):
            try:
                dd = datetime.strptime(str(dd)[:10], '%Y-%m-%d')
            except Exception:
                continue
        if dd.year == year and dd.month == month:
            results.append(data)
    return results


def get_all_preview(year, month):
    """获取3个项目指定月份的报价单预览数据"""
    result = {}
    for key, cfg in PROJECTS.items():
        history_dir = QUOTE_BASE / cfg['dir']
        if key == 'liandishenkong':
            x3_quotes = []
            for f in scan_quote_files(history_dir):
                try:
                    data = read_x3_quote_by_type(f, cfg['delivery_cell'])
                except Exception:
                    continue
                dd = data['delivery_date']
                if not isinstance(dd, datetime):
                    try:
                        dd = datetime.strptime(str(dd)[:10], '%Y-%m-%d')
                    except Exception:
                        continue
                if dd.year == year and dd.month == month:
                    x3_quotes.append(data)
            total_words = sum(sum(tw.values()) for tw in (q['type_words'] for q in x3_quotes))
            total_amount = 0
            for q in x3_quotes:
                for tn, w in q['type_words'].items():
                    total_amount += w * X3_PRICES[tn]
            quote_list = []
            for q in x3_quotes:
                words = sum(q['type_words'].values())
                amount = sum(w * X3_PRICES[tn] for tn, w in q['type_words'].items())
                quote_list.append({
                    'file_name': q['file_name'],
                    'delivery_date': q['delivery_date'].strftime('%Y-%m-%d') if isinstance(q['delivery_date'], datetime) else str(q['delivery_date']),
                    'total_words': round(words, 1),
                    'total_amount_pretax': round(amount, 2),
                    'total_amount': round(amount * 1.06, 2),
                    'type_words': {k: round(v, 1) for k, v in q['type_words'].items()},
                })
            result[key] = {
                'display': cfg['display'],
                'quotes': quote_list,
                'total_words': round(total_words, 1),
                'total_amount_pretax': round(total_amount, 2),
                'total_amount': round(total_amount * 1.06, 2),
            }
        else:
            quotes = find_quotes_for_month(
                history_dir, cfg['word_col'], cfg['amount_col'], cfg['delivery_cell'], year, month
            )
            result[key] = {
                'display': cfg['display'],
                'quotes': [{
                    'file_name': q['file_name'],
                    'delivery_date': q['delivery_date'].strftime('%Y-%m-%d') if isinstance(q['delivery_date'], datetime) else str(q['delivery_date']),
                    'total_words': round(q['total_words'], 1),
                    'total_amount_pretax': round(q['total_amount'], 2),
                    'total_amount': round(q['total_amount'] * 1.06, 2),
                } for q in quotes],
                'total_words': round(sum(q['total_words'] for q in quotes), 1),
                'total_amount_pretax': round(sum(q['total_amount'] for q in quotes), 2),
                'total_amount': round(sum(q['total_amount'] for q in quotes) * 1.06, 2),
            }
    return result


def _generate_x3_excel(quotes_by_type, year, month, output_dir):
    """生成X3结算单：按批次+类型分行"""
    cfg = PROJECTS['liandishenkong']
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"【大连游者之家】结算单_《{cfg['short']}》{year}年{month}月.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime('%H%M%S')
            output_path = output_dir / f"【大连游者之家】结算单_《{cfg['short']}》{year}年{month}月_{ts}.xlsx"

    shutil.copy2(str(TEMPLATE_PATH), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    today = datetime.now()
    ws['C7'] = f'结算月：{year}年{month}月'
    ws['I7'] = today

    data_start_row = 10
    max_data_rows = 14
    row = data_start_row

    for q in quotes_by_type:
        types = q['type_words']
        is_single_tep = len(types) == 1 and "TEP" in types
        for type_name, words in types.items():
            if row >= data_start_row + max_data_rows:
                break
            price = X3_PRICES[type_name]
            suffix = "" if is_single_tep else type_name
            ws.cell(row=row, column=3, value=q['file_name'] + suffix)
            ws.cell(row=row, column=6, value=f'{month}月')
            ws.cell(row=row, column=7, value=round(words, 1))
            ws.cell(row=row, column=8, value=price)
            row += 1

    total_row = data_start_row + max_data_rows
    for r in range(row, total_row):
        ws.row_dimensions[r].hidden = True

    wb.save(str(output_path))
    return output_path


def _generate_single_excel(project_key, quotes, year, month, output_dir):
    cfg = PROJECTS[project_key]
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"【大连游者之家】结算单_《{cfg['short']}》{year}年{month}月.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime('%H%M%S')
            output_path = output_dir / f"【大连游者之家】结算单_《{cfg['short']}》{year}年{month}月_{ts}.xlsx"

    shutil.copy2(str(TEMPLATE_PATH), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    today = datetime.now()
    ws['C7'] = f'结算月：{year}年{month}月'
    ws['I7'] = today

    data_start_row = 10
    max_data_rows = 14

    for i, q in enumerate(quotes):
        if i >= max_data_rows:
            break
        r = data_start_row + i
        ws.cell(row=r, column=3, value=q['file_name'])
        ws.cell(row=r, column=6, value=f'{month}月')
        ws.cell(row=r, column=7, value=q['total_words'])
        if q['total_words'] > 0:
            ws.cell(row=r, column=8, value=round(q['total_amount'] / q['total_words'], 4))
        else:
            ws.cell(row=r, column=8, value=0)

    total_row = data_start_row + max_data_rows
    last_data_row = data_start_row + min(len(quotes), max_data_rows) - 1
    for r in range(last_data_row + 1, total_row):
        ws.row_dimensions[r].hidden = True

    wb.save(str(output_path))
    return output_path


def generate_all(year, month):
    """生成3个项目结算单，保存到统一目录"""
    output_dir = OUTPUT_BASE / f"{year}年{month}月" / "叠纸"
    results = {}
    all_quotes = {}
    for key, cfg in PROJECTS.items():
        history_dir = QUOTE_BASE / cfg['dir']
        if key == 'liandishenkong':
            # X3: 按类型分类读取
            x3_quotes = []
            for f in scan_quote_files(history_dir):
                try:
                    data = read_x3_quote_by_type(f, cfg['delivery_cell'])
                except Exception as e:
                    print(f'[WARN] 读取报价单失败: {f.name}: {e}')
                    continue
                dd = data['delivery_date']
                if not isinstance(dd, datetime):
                    try:
                        dd = datetime.strptime(str(dd)[:10], '%Y-%m-%d')
                    except Exception:
                        continue
                if dd.year == year and dd.month == month:
                    x3_quotes.append(data)
            all_quotes[key] = x3_quotes
            if not x3_quotes:
                results[key] = {'path': None, 'amount': 0, 'words': 0}
                continue
            path = _generate_x3_excel(x3_quotes, year, month, output_dir)
            total_words = sum(sum(tw.values()) for tw in (q['type_words'] for q in x3_quotes))
            total_amount = 0
            for q in x3_quotes:
                for tn, w in q['type_words'].items():
                    total_amount += w * X3_PRICES[tn]
            results[key] = {'path': str(path), 'amount': round(total_amount * 1.06, 2), 'words': round(total_words, 1)}
        else:
            quotes = find_quotes_for_month(
                history_dir, cfg['word_col'], cfg['amount_col'], cfg['delivery_cell'], year, month
            )
            all_quotes[key] = quotes
            if not quotes:
                results[key] = {'path': None, 'amount': 0, 'words': 0}
                continue
            path = _generate_single_excel(key, quotes, year, month, output_dir)
            total_amount = round(sum(q['total_amount'] for q in quotes) * 1.06, 2)
            total_words = round(sum(q['total_words'] for q in quotes), 1)
            results[key] = {'path': str(path), 'amount': total_amount, 'words': total_words}

    # 用Excel重算所有生成的结算单
    import pythoncom
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        for key, info in results.items():
            if info['path']:
                wb = excel.Workbooks.Open(info['path'])
                excel.CalculateUntilAsyncQueriesDone()
                wb.Save()
                wb.Close()
        excel.Quit()
    except Exception as e:
        print(f'[WARN] 结算单Excel重算失败: {e}')
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    # 生成开票信息
    generate_invoice_info(results, year, month)

    # 移动已结算报价单到已结算目录
    for key, cfg in PROJECTS.items():
        quotes = all_quotes.get(key, [])
        if not quotes:
            continue
        history_dir = QUOTE_BASE / cfg['dir']
        settled_dir = history_dir / "已结算" / f"{year}年{month}月"
        settled_dir.mkdir(parents=True, exist_ok=True)
        for q in quotes:
            src = history_dir / q['file_name']
            if src.exists():
                shutil.move(str(src), str(settled_dir / q['file_name']))

    return results


def generate_invoice_info(results, year, month):
    """生成开票信息文本文件"""
    output_dir = OUTPUT_BASE / f"{year}年{month}月" / "叠纸"
    output_dir.mkdir(parents=True, exist_ok=True)
    inv_path = output_dir / "开票信息.txt"

    lines = [f"开票信息 —— {year}年{month}月", ""]

    order = ['shining_nikki', 'niki_xinzuo', 'liandishenkong']
    for key in order:
        info = results.get(key) or {}
        amount = info.get('amount', 0)
        company = INVOICE_COMPANIES[key]
        lines.append(f"【{company['label']}】")
        lines.append(f"单位名称：{company['name']}")
        lines.append(f"纳税人识别号：{company['tax_id']}")
        lines.append(f"开户行：{company['bank']}")
        lines.append(f"银行账号：{company['account']}")
        lines.append(f"单位地址：{company['address']}")
        lines.append(f"电话：{company['phone']}")
        lines.append(f"翻译费：{amount:.2f}")
        lines.append("")

    inv_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f'[OK] 开票信息已保存: {inv_path}')
    return inv_path


def get_invoice_info(year, month):
    """读取已有的开票信息"""
    inv_path = OUTPUT_BASE / f"{year}年{month}月" / "叠纸" / "开票信息.txt"
    if inv_path.exists():
        return inv_path.read_text(encoding='utf-8')
    return None


def main():
    import argparse
    parser = argparse.ArgumentParser(description='叠纸3项目月度结算单生成')
    parser.add_argument('--year', type=int, default=2026)
    parser.add_argument('--month', type=int, default=5)
    parser.add_argument('--project', type=str, default=None)
    args = parser.parse_args()

    if args.project and args.project in PROJECTS:
        cfg = PROJECTS[args.project]
        history_dir = QUOTE_BASE / cfg['dir']
        quotes = find_quotes_for_month(
            history_dir, cfg['word_col'], cfg['amount_col'], cfg['delivery_cell'],
            args.year, args.month
        )
        print(f'找到 {len(quotes)} 份报价单')
        if quotes:
            output_dir = OUTPUT_BASE / f"{args.year}年{args.month}月" / "叠纸"
            _generate_single_excel(args.project, quotes, args.year, args.month, output_dir)
    else:
        generate_all(args.year, args.month)

    print('完成！')


if __name__ == '__main__':
    main()
