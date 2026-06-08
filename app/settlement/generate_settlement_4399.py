"""4399 月度结算：从飞书读取 → 结算单 Excel/PDF + Invoice PDF"""
from __future__ import annotations

import calendar
import json
import os
import shutil
import sys
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from quote_system.feishu_client import FeishuClient
from quote_system.paths import get_settlement_dir

ROOT = Path(__file__).resolve().parent.parent.parent
TEMPLATE_DIR = ROOT / "app" / "模板" / "结算模板"

SETTLEMENT_TEMPLATE = "01【结算单】-2026年4月份翻译费用4399模板.xlsx"
INVOICE_TEMPLATE = "01Invoice-2026年4月份翻译费用4399模板.xlsx"

SPREADSHEET_TOKEN_4399 = "YcKKsgMtUhrghlt65ubczJ21n9c"
SHEET_ID_4399 = "a63cb8"

SPREADSHEET_TOKEN_BOQI = "EmtEs2ZtAhC2VqtlTSAcYVm8ntc"
SHEET_ID_BOQI = "31760b"

UNIT_PRICE = 0.52
DATA_START_ROW = 16
DATA_END_ROW = 40

CONFIG_DIR = ROOT / "app" / "quote_system" / "config"
PROJECT_CODE_FILE = CONFIG_DIR / "4399_projects.json"

DEFAULT_PROJECT_CODES = {
    "冒险大作战": "01",
    "明日特工队": "02",
    "指尖无双": "03",
    "主宰世界": "04",
    "波奇": "05",
}


def _load_project_codes() -> dict[str, str]:
    if PROJECT_CODE_FILE.exists():
        try:
            data = json.loads(PROJECT_CODE_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return dict(DEFAULT_PROJECT_CODES)


def _save_project_codes(codes: dict[str, str]):
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    PROJECT_CODE_FILE.write_text(json.dumps(codes, ensure_ascii=False, indent=2), encoding="utf-8")


def _get_project_code(project_name: str) -> str:
    codes = _load_project_codes()
    if project_name in codes:
        return codes[project_name]
    used = {int(v) for v in codes.values() if v.isdigit()}
    next_code = 1
    while next_code in used:
        next_code += 1
    code = f"{next_code:02d}"
    codes[project_name] = code
    _save_project_codes(codes)
    return code


def get_first_workday(year: int, month: int) -> datetime:
    d = datetime(year, month, 1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def get_exchange_rate(year: int, month: int) -> float | None:
    d = datetime(year, month, 1)
    max_attempts = 10
    for _ in range(max_attempts):
        if d.weekday() < 5:
            rate = _fetch_rate_for_date(d.strftime("%Y-%m-%d"))
            if rate is not None:
                return rate
        d += timedelta(days=1)
    return None


def _fetch_rate_for_date(date_str: str) -> float | None:
    url = f"http://www.safe.gov.cn/AppStructured/hlw/jsonRmb.do?date={date_str}"
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "http://www.safe.gov.cn/safe/rmbhlzjj/index.html",
        })
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode())
        for item in data:
            if item[1] == "美元":
                return float(item[2]) / 100
        return None
    except Exception:
        return None


def excel_serial_to_date(serial) -> datetime | None:
    try:
        num = float(serial)
    except (TypeError, ValueError):
        return None
    base = datetime(1899, 12, 30)
    try:
        return base + timedelta(days=num)
    except Exception:
        return None


def read_feishu_data_4399(year: int, month: int) -> dict[str, list[dict]]:
    client = FeishuClient(sheet_id=SHEET_ID_4399, spreadsheet_token=SPREADSHEET_TOKEN_4399)
    rows = client.read_sheet()

    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)
    base = datetime(1899, 12, 30)
    serial_start = (month_start - base).days
    serial_end = (month_end - base).days

    projects: dict[str, list[dict]] = {}
    for i, row in enumerate(rows):
        if i == 0:
            continue

        project_name = str(row[2] or "").strip() if len(row) > 2 else ""
        if not project_name:
            continue

        deliv_val = row[4] if len(row) > 4 else None
        if deliv_val is None:
            continue
        try:
            deliv_serial = float(deliv_val)
            if not (serial_start <= deliv_serial < serial_end):
                continue
        except (TypeError, ValueError):
            continue

        req_name = str(row[1] or "").strip() if len(row) > 1 else ""
        deliv_date = deliv_serial

        k_val = float(row[10]) if len(row) > 10 and row[10] is not None else 0
        l_val = float(row[11]) if len(row) > 11 and row[11] is not None else 0
        m_val = float(row[12]) if len(row) > 12 and row[12] is not None else 0
        word_count = round(k_val - m_val + (l_val - k_val) * 0.25)

        records = projects.setdefault(project_name, [])
        records.append({
            "req_name": req_name,
            "deliv_date": deliv_date,
            "word_count": word_count,
            "unit_price": UNIT_PRICE,
            "row_index": i + 1,
        })

    return projects


def read_feishu_data_boqi(year: int, month: int) -> list[dict]:
    client = FeishuClient(sheet_id=SHEET_ID_BOQI, spreadsheet_token=SPREADSHEET_TOKEN_BOQI)
    rows = client.read_sheet()

    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)
    base = datetime(1899, 12, 30)
    serial_start = (month_start - base).days
    serial_end = (month_end - base).days

    records = []
    for i, row in enumerate(rows):
        if i == 0:
            continue

        deliv_val = row[3] if len(row) > 3 else None
        if deliv_val is None:
            continue
        try:
            deliv_serial = float(deliv_val)
            if not (serial_start <= deliv_serial < serial_end):
                continue
        except (TypeError, ValueError):
            continue

        status = str(row[6] or "").strip() if len(row) > 6 else ""
        if status != "已交付":
            continue

        amount = float(row[5]) if len(row) > 5 and row[5] is not None else 0
        word_count = round(amount / UNIT_PRICE, 2)

        records.append({
            "req_name": str(row[1] or "").strip() if len(row) > 1 else "",
            "deliv_date": deliv_serial,
            "word_count": word_count,
            "unit_price": UNIT_PRICE,
            "row_index": i + 1,
        })

    return records


def generate_settlement_excel(records: list[dict], project_name: str, year: int, month: int,
                              output_dir: Path, exchange_rate: float) -> Path:
    template = TEMPLATE_DIR / SETTLEMENT_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"找不到模板: {template}")

    project_code = _get_project_code(project_name)
    filename = f"{project_code}【结算单】-{year}年{month}月份翻译费用{project_name}.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime("%H%M%S")
            filename = f"{project_code}【结算单】-{year}年{month}月份翻译费用{project_name}_{ts}.xlsx"
            output_path = output_dir / filename

    shutil.copy2(str(template), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    today = datetime.now()
    _, last_day = calendar.monthrange(year, month)

    ws["D9"] = f"{today.strftime('%Y%m%d')}-{project_code}"
    ws["D13"] = f"《{project_name}》"
    ws["D14"] = f"{year}/{month}/1-{year}/{month}/{last_day}"

    ws["D46"] = exchange_rate
    ws["H3"] = today

    max_data_rows = DATA_END_ROW - DATA_START_ROW + 1
    if len(records) > max_data_rows:
        extra = len(records) - max_data_rows
        for _ in range(extra):
            ws.insert_rows(DATA_END_ROW + 1)
        for r in range(DATA_START_ROW + len(records), DATA_START_ROW + len(records) + extra + 50):
            pass

    for i, rec in enumerate(records):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=3, value=rec["req_name"])
        dt = excel_serial_to_date(rec["deliv_date"])
        if dt:
            ws.cell(row=r, column=4, value=dt)
            ws.cell(row=r, column=4).number_format = "m/d"
        ws.cell(row=r, column=5, value=rec["word_count"])
        ws.cell(row=r, column=6, value=UNIT_PRICE)

    last_data_row = DATA_START_ROW + len(records) - 1
    for r in range(last_data_row + 1, DATA_END_ROW + 1):
        ws.row_dimensions[r].hidden = True

    wb.save(str(output_path))
    return output_path


def generate_invoice_excel(settlement_xlsx_path: Path, project_name: str, year: int, month: int,
                           output_dir: Path) -> Path:
    template = TEMPLATE_DIR / INVOICE_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"找不到模板: {template}")

    project_code = _get_project_code(project_name)
    filename = f"{project_code}Invoice-{year}年{month}月份翻译费用{project_name}.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime("%H%M%S")
            filename = f"{project_code}Invoice-{year}年{month}月份翻译费用{project_name}_{ts}.xlsx"
            output_path = output_dir / filename

    shutil.copy2(str(template), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    today = datetime.now()
    _, last_day = calendar.monthrange(year, month)

    ws["H11"] = today
    ws["D12"] = f"{today.strftime('%Y%m%d')}-{project_code}"

    start_str = f"{year}/{month}/1"
    end_str = f"{year}/{month}/{last_day}"
    ws["B17"] = f"结账月份：{start_str} ~ {end_str}"

    ws["C19"] = f"{project_name}本地化翻译服务"

    usd_amount = _read_settlement_h42(settlement_xlsx_path)
    ws["G19"] = usd_amount

    wb.save(str(output_path))
    return output_path


def _read_settlement_h42(xlsx_path: Path) -> float:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        val = ws["H42"].value
        wb.close()
        if isinstance(val, (int, float)) and val > 0:
            return round(float(val), 2)
    except Exception:
        pass

    _recalc_excel(xlsx_path)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        val = ws["H42"].value
        wb.close()
        if isinstance(val, (int, float)):
            return round(float(val), 2)
    except Exception:
        pass

    raise ValueError(f"无法从结算单读取 H42 合计金额: {xlsx_path}")


def _recalc_excel(file_path: Path):
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(file_path.resolve()))
        excel.CalculateUntilAsyncQueriesDone()
        wb.Save()
        wb.Close()
    finally:
        try:
            if wb:
                wb.Close()
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _convert_to_pdf(xlsx_path: Path) -> Path:
    pdf_path = xlsx_path.with_suffix(".pdf")
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        wb.SaveAs(str(pdf_path.resolve()), FileFormat=57)
        wb.Close()
    finally:
        try:
            if wb:
                wb.Close()
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
    return pdf_path


def generate_all(year: int, month: int, exchange_rate: float) -> dict:
    output_dir = get_settlement_dir() / f"{year}年{month}月" / "4399"
    output_dir.mkdir(parents=True, exist_ok=True)

    result: dict[str, dict] = {}

    main_data = read_feishu_data_4399(year, month)
    for project_name, records in main_data.items():
        if not records:
            continue
        _process_project(records, project_name, year, month, output_dir, exchange_rate, result)

    boqi_records = read_feishu_data_boqi(year, month)
    if boqi_records:
        _process_project(boqi_records, "波奇", year, month, output_dir, exchange_rate, result)

    _generate_summary_excel(result, year, month, output_dir)

    return result


def _generate_summary_excel(result: dict, year: int, month: int, output_dir: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "结算汇总"

    group1 = ["指尖无双", "主宰世界"]
    group2 = ["冒险大作战", "明日特工队", "波奇"]

    bold = openpyxl.styles.Font(bold=True)

    row = 1
    ws.cell(row=row, column=1, value="项目名").font = bold
    ws.cell(row=row, column=2, value="金额(USD)").font = bold
    row += 1
    for name in group1:
        if name in result:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=result[name]["usd_amount"])
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
    sub1 = sum(result[n]["usd_amount"] for n in group1 if n in result)
    ws.cell(row=row, column=1, value="合计").font = bold
    ws.cell(row=row, column=2, value=sub1).font = bold
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    row += 2

    ws.cell(row=row, column=1, value="项目名").font = bold
    ws.cell(row=row, column=2, value="金额(USD)").font = bold
    row += 1
    for name in group2:
        if name in result:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=result[name]["usd_amount"])
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
    sub2 = sum(result[n]["usd_amount"] for n in group2 if n in result)
    ws.cell(row=row, column=1, value="合计").font = bold
    ws.cell(row=row, column=2, value=sub2).font = bold
    ws.cell(row=row, column=2).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15

    path = output_dir / "结算汇总.xlsx"
    wb.save(str(path))
    print(f"  [OK] 结算汇总: {path.name}")


def _process_project(records: list[dict], project_name: str, year: int, month: int,
                     output_dir: Path, exchange_rate: float, result: dict):
    print(f"  生成 {project_name} 结算单 ({len(records)} 条)...")
    settlement_xlsx = generate_settlement_excel(records, project_name, year, month, output_dir, exchange_rate)

    print(f"  生成 {project_name} Invoice...")
    invoice_xlsx = generate_invoice_excel(settlement_xlsx, project_name, year, month, output_dir)

    usd_amount = _read_settlement_h42(settlement_xlsx)

    print(f"  转换 {project_name} PDF...")
    settlement_pdf = _convert_to_pdf(settlement_xlsx)
    invoice_pdf = _convert_to_pdf(invoice_xlsx)

    os.remove(str(invoice_xlsx))

    result[project_name] = {
        "xlsx": str(settlement_xlsx),
        "settlement_pdf": str(settlement_pdf),
        "invoice_pdf": str(invoice_pdf),
        "usd_amount": usd_amount,
    }
    print(f"  [OK] {project_name}: USD {usd_amount:,.2f}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="4399 月度结算")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--month", type=int, default=5)
    parser.add_argument("--rate", type=float, required=True, help="人民币对美元汇率")
    args = parser.parse_args()

    result = generate_all(args.year, args.month, args.rate)

    print("\n=== 汇总 ===")
    table1 = []  # 指尖无双 + 主宰世界
    table2 = []  # 其他
    for name, info in result.items():
        if name in ("指尖无双", "主宰世界"):
            table1.append((name, info["usd_amount"]))
        else:
            table2.append((name, info["usd_amount"]))

    if table1:
        print("\n【表1】指尖无双 + 主宰世界")
        for name, amt in table1:
            print(f"  {name}  USD {amt:,.2f}")
    if table2:
        print("\n【表2】其他项目")
        for name, amt in table2:
            print(f"  {name}  USD {amt:,.2f}")


if __name__ == "__main__":
    main()
