"""战双版更月度结算单生成"""
import sys
import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
import re

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from quote_system.feishu_client import FeishuClient
from quote_system.paths import get_settlement_dir

SPREADSHEET_TOKEN = "Wup0wnUPIiiIr2k8T23c4zASnjd"
SHEET_ID = "WmpiFq"
UNIT_PRICE = 0.64

BASE_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATE_DIR = BASE_DIR / "app" / "模板" / "结算模板"


def excel_serial_to_date(serial):
    base = datetime(1899, 12, 30)
    return base + timedelta(days=float(serial))


def read_feishu_data(year, month):
    client = FeishuClient(spreadsheet_token=SPREADSHEET_TOKEN, sheet_id=SHEET_ID)
    data = client.read_sheet()

    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)
    base = datetime(1899, 12, 30)
    serial_start = (month_start - base).days
    serial_end = (month_end - base).days

    records = []
    for i, row in enumerate(data):
        if i == 0:
            continue
        if len(row) < 9:
            continue

        source = str(row[1] or "").strip()
        if source != "邮件":
            continue

        status = str(row[5] or "").strip()
        if status != "已报价":
            continue

        deliv_val = row[3] if len(row) > 3 else None
        if deliv_val is None:
            continue
        try:
            deliv_serial = float(deliv_val)
            if not (serial_start <= deliv_serial < serial_end):
                continue
        except (TypeError, ValueError):
            continue

        word_count = row[8] if len(row) > 8 else None
        records.append({
            'row_index': i + 1,
            'req_name': str(row[0] or ""),
            'deliv_date': deliv_serial,
            'word_count': float(word_count) if word_count is not None else 0,
            'file_name': str(row[7] or "") if len(row) > 7 else "",
        })

    return records


def generate_settlement_excel(records, year, month, output_dir):
    template = TEMPLATE_DIR / "【大连游者之家】结算单_战双版更需求26年5月模板.xlsx"
    filename = f"【大连游者之家】结算单_战双版更需求{year}年{month}月.xlsx"
    output_path = output_dir / filename

    if output_path.exists():
        try:
            os.remove(str(output_path))
        except PermissionError:
            ts = datetime.now().strftime('%H%M%S')
            filename = f"【大连游者之家】结算单_战双版更需求{year}年{month}月_{ts}.xlsx"
            output_path = output_dir / filename

    shutil.copy2(str(template), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    today = datetime.now()
    ws['C5'] = int(today.strftime('%Y%m%d'))
    ws['G5'] = today

    def _sort_key(rec):
        name = rec['req_name']
        m = re.search(r'(\d+(?:\.\d+)?)[^\d]*(\d+)$', name)
        if m:
            return (tuple(int(x) for x in m.group(1).split('.')), int(m.group(2)))
        return (name,)
    records.sort(key=_sort_key)

    data_start_row = 13

    # 填充数据
    for i, rec in enumerate(records):
        r = data_start_row + i
        ws.cell(row=r, column=2, value=rec['file_name'])   # B: 报价单文件名
        ws.cell(row=r, column=4, value=rec['deliv_date'])
        ws.cell(row=r, column=4).number_format = 'm/d'
        if rec['word_count']:
            ws.cell(row=r, column=5, value=rec['word_count'])
        ws.cell(row=r, column=6, value=UNIT_PRICE)
        ws.cell(row=r, column=10, value=rec['req_name'])   # J: 委托需求

    # 隐藏最后一行数据到合计行之间的空行
    last_data_row = data_start_row + len(records) - 1
    # 找合计行
    total_row = None
    for row in range(last_data_row + 1, ws.max_row + 1):
        for col in range(1, 19):
            val = ws.cell(row=row, column=col).value
            if val and '合' in str(val):
                total_row = row
                break
        if total_row:
            break
    if total_row:
        for r in range(last_data_row + 1, total_row):
            ws.row_dimensions[r].hidden = True

    wb.save(str(output_path))
    print(f'[OK] 结算单已保存: {output_path}')
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description='战双版更月度结算单生成')
    parser.add_argument('--year', type=int, default=2026)
    parser.add_argument('--month', type=int, default=5)
    parser.add_argument('--output', type=str, default=None)
    args = parser.parse_args()

    year, month = args.year, args.month
    output_dir = Path(args.output) if args.output else get_settlement_dir() / f"{year}年{month}月" / "库洛游戏" / "战双版更"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f'读取飞书数据: {year}年{month}月...')
    records = read_feishu_data(year, month)
    print(f'找到 {len(records)} 条交付记录')

    if not records:
        print('没有找到匹配的交付记录，退出。')
        return

    generate_settlement_excel(records, year, month, output_dir)
    print('完成！')


if __name__ == '__main__':
    main()
