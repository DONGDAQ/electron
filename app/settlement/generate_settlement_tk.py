"""TK 结算：对账单(OrderList) → 结算单PDF + Invoice PDF。"""
from __future__ import annotations

import calendar
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from quote_system.paths import get_settlement_dir

TEMPLATE_DIR = ROOT / "模板" / "结算模板"

ORDERLIST_TEMPLATE = "【Gamer House】【Trickcal】订单明细表OrderList_2026年3月模板.xlsx"
SETTLEMENT_TEMPLATE = "【Trickcal】翻译费用明细及结算清单2026年3月模板.xlsx"
INVOICE_TEMPLATE = "Invoice-【Trickcal】2026年3月翻译费用模板.xlsx"

PROJECT_NAME = "TK"
COMPANY = "TK"


def copy_orderlist(year: int, month: int) -> Path:
    """复制订单明细表模板到结算文件夹并返回路径。"""
    template = TEMPLATE_DIR / ORDERLIST_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"找不到模板: {template}")

    output_dir = get_settlement_dir() / f"{year}年{month}月" / COMPANY
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"【Gamer House】【Trickcal】订单明细表OrderList_{year}年{month}月.xlsx"
    output_path = output_dir / filename
    if not output_path.exists():
        shutil.copy2(str(template), str(output_path))

    return output_path


def _read_orderlist_data(orderlist_path: Path) -> tuple[float, int]:
    """从订单明细表读取 M2 单元格(总金额)和 J 列总字数。"""
    wb = openpyxl.load_workbook(orderlist_path, data_only=True)
    ws = wb["KOMOE"] if "KOMOE" in wb.sheetnames else wb.active

    m2_val = ws.cell(row=2, column=13).value
    if not isinstance(m2_val, (int, float)) or m2_val <= 0:
        # M2 是公式，缓存为空，用 Excel 重算
        _recalc_excel(orderlist_path)
        wb = openpyxl.load_workbook(orderlist_path, data_only=True)
        ws = wb["KOMOE"] if "KOMOE" in wb.sheetnames else wb.active
        m2_val = ws.cell(row=2, column=13).value

    if not isinstance(m2_val, (int, float)) or m2_val <= 0:
        raise ValueError(f"无法读取订单明细表 M2 总金额: {m2_val}")

    total_j = 0
    for r in range(3, ws.max_row + 1):
        j_val = ws.cell(row=r, column=10).value
        if isinstance(j_val, (int, float)):
            total_j += int(j_val)
        elif j_val is None and _is_empty_row(ws, r):
            break

    from decimal import Decimal, ROUND_HALF_UP
    amount = float(Decimal(str(m2_val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    return amount, total_j


def _is_empty_row(ws, r: int) -> bool:
    for c in range(1, 11):
        if ws.cell(row=r, column=c).value is not None:
            return False
    return True


def _recalc_excel(file_path: Path):
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(file_path.resolve()))
        wb.Save()
        wb.Close()
    finally:
        try:
            if wb: wb.Close()
        except Exception: pass
        try:
            excel.Quit()
        except Exception: pass
        try:
            pythoncom.CoUninitialize()
        except Exception: pass


def generate_settlement_pdf(year: int, month: int) -> Path:
    """生成结算单 PDF。"""
    orderlist_path = _find_orderlist(year, month)
    total_amount, total_words = _read_orderlist_data(orderlist_path)

    template = TEMPLATE_DIR / SETTLEMENT_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"找不到模板: {template}")

    output_dir = get_settlement_dir() / f"{year}年{month}月" / COMPANY
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f"【Trickcal】翻译费用明细及结算清单{year}年{month}月.xlsx"
    pdf_path = xlsx_path.with_suffix(".pdf")
    if not xlsx_path.exists():
        shutil.copy2(str(template), str(xlsx_path))
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    _, last_day = calendar.monthrange(year, month)
    ws["C9"] = f"TK{month}月份需求（详见订单明细表）"
    ws["D9"] = "翻译/Translation"
    ws["E9"] = f"{year}/{month}/1"
    ws["F9"] = f"{year}/{month}/{last_day}"
    ws["G9"] = total_words
    ws["I9"] = total_amount

    today = datetime.now()
    ws["B19"] = (
        f"乙方：大连游者之家信息技术有限公司\n"
        f"负责人：董大千\n"
        f"电话：13940868702\n"
        f"日期：【{today.year}】年【{today.month}】月【{today.day}】日"
    )

    wb.save(str(xlsx_path))

    _convert_xlsx_to_pdf(xlsx_path, pdf_path)
    # xlsx_path.unlink()  # 暂时保留Excel文件方便检查格式

    return pdf_path


def generate_invoice_pdf(year: int, month: int) -> Path:
    """生成 Invoice PDF。"""
    orderlist_path = _find_orderlist(year, month)
    total_amount, _ = _read_orderlist_data(orderlist_path)

    template = TEMPLATE_DIR / INVOICE_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"找不到模板: {template}")

    output_dir = get_settlement_dir() / f"{year}年{month}月" / COMPANY
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f"Invoice-【Trickcal】{year}年{month}月翻译费用.xlsx"
    pdf_path = xlsx_path.with_suffix(".pdf")
    if not xlsx_path.exists():
        shutil.copy2(str(template), str(xlsx_path))
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["invoice"] if "invoice" in wb.sheetnames else wb.active

    today = datetime.now()
    _, last_day = calendar.monthrange(year, month)

    ws["K11"] = today.strftime("%d-%b-%y")
    ws["D11"] = f"HK-{today.strftime('%Y%m%d')}"
    ws["B16"] = f"结账月份：{year}/{month}/1~{year}/{month}/{last_day}"
    ws["J18"] = total_amount
    ws["K18"] = total_amount

    wb.save(str(xlsx_path))

    _convert_xlsx_to_pdf(xlsx_path, pdf_path)
    # xlsx_path.unlink()  # 暂时保留Excel文件方便检查格式

    return pdf_path


def _find_orderlist(year: int, month: int) -> Path:
    output_dir = get_settlement_dir() / f"{year}年{month}月" / COMPANY
    if not output_dir.exists():
        raise FileNotFoundError(f"结算目录不存在: {output_dir}（请先生成对账单）")

    pattern = f"【Gamer House】【Trickcal】订单明细表OrderList_{year}年{month}月.xlsx"
    path = output_dir / pattern
    if path.exists():
        return path

    matches = sorted(output_dir.glob(f"【Gamer House】【Trickcal】订单明细表OrderList_{year}年{month}月*.xlsx"))
    if matches:
        return matches[0]

    raise FileNotFoundError(f"找不到订单明细表: {pattern}")


def _convert_xlsx_to_pdf(xlsx_path: Path, pdf_path: Path) -> None:
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        wb.SaveAs(str(pdf_path.resolve()), FileFormat=57)
        wb.Close()
    finally:
        if wb:
            try: wb.Close()
            except Exception: pass
        excel.Quit()
        try:
            pythoncom.CoUninitialize()
        except Exception: pass
